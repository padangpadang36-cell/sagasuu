# -*- coding: utf-8 -*-
"""依頼一覧Excelに管理番号列を追加する"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from pathlib import Path

path = Path(r'C:\Users\shimi\Desktop\syataku.com\入力データ\依頼一覧_テスト.xlsx')
wb = openpyxl.load_workbook(path)
ws = wb.active

# 既に管理番号列があるか確認
headers = [cell.value for cell in ws[1]]
print("現在の列:", headers)

if '管理番号' in headers:
    print("管理番号列は既に存在します")
else:
    # 先頭に管理番号列を挿入
    ws.insert_cols(1)
    ws.cell(1, 1, '管理番号')

    # データ行にA001, A002...を入れる（例行はスキップ）
    counter = 1
    for row_idx in range(2, ws.max_row + 1):
        name_val = ws.cell(row_idx, 2).value  # 氏名は列2になった
        if name_val and not str(name_val).startswith('例：') and not str(name_val).startswith('↑'):
            ws.cell(row_idx, 1, f'A{counter:03d}')
            counter += 1

    wb.save(path)
    print("管理番号列を追加して保存しました")

    # 確認
    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2.active
    print("更新後の列:", [cell.value for cell in ws2[1]])
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if any(row):
            print("  行:", row[:4])
