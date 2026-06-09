import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

BASE_DIR = Path(__file__).parent

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "依頼一覧"

headers = [
    "氏名", "転勤元住所", "勤務先住所（転勤先）",
    "希望エリア", "入居希望日", "希望社宅種類",
    "家賃上限（円）", "希望間取り", "通勤方法", "通勤可能時間"
]

header_fill = PatternFill("solid", fgColor="1a237e")
header_font = Font(bold=True, color="FFFFFF", name="メイリオ")

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

test_row = [
    "テスト太郎",
    "東京都新宿区西新宿2-8-1",
    "大阪府大阪市北区梅田3-1-1",
    "大阪市北区",
    "2026/06/01",
    "家具家電なし社宅",
    80000,
    "1K",
    "電車",
    "30分以内",
]
ws.append(test_row)

col_widths = [15, 30, 30, 15, 15, 20, 15, 12, 12, 12]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

out_path = BASE_DIR / "test_data.xlsx"
wb.save(out_path)
print(f"テストデータ作成完了: {out_path}")
