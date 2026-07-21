"""
社宅物件提案 自動化スクリプト
- ATBBにログインして物件を検索
- Googleマップで通勤ルートのスクリーンショットを取得
- PDF形式で提案書を出力
"""
import sys
import io
import asyncio
import os
from pathlib import Path
from dotenv import load_dotenv
from datetime import datetime

# Windows コンソールの文字化け対策
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import re
import shutil
import openpyxl
from playwright.async_api import async_playwright
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors
from PIL import Image
from pypdf import PdfReader, PdfWriter

# ─── パス設定 ───────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent.parent   # プロジェクトルート
INPUT_DIR   = BASE_DIR / "入力データ"
OUTPUT_DIR  = BASE_DIR / "出力PDF"
SHOT_DIR    = BASE_DIR / "システムファイル" / "デバッグ" / "screenshots"
DETAIL_DIR  = BASE_DIR / "出力PDF" / "物件詳細"

load_dotenv(BASE_DIR / ".env")

# ─── ATBB 認証情報 ─────────────────────────────────────────
ATBB_URL  = "https://atbb.athome.jp/"
ATBB_ID   = os.getenv("ATBB_ID", "")
ATBB_PASS = os.getenv("ATBB_PASS", "")

# ─── 東建ルームサーチ 認証情報 ──────────────────────────────
HM_URL  = "https://www.homemate.co.jp/hmroom/"
HM_ID   = os.getenv("HM_ID", "")
HM_PASS = os.getenv("HM_PASS", "")

# ─── D-Room (大和リビング) 認証情報 ────────────────────────
DROOM_LOGIN_URL    = "https://anavi.daiwaliving.co.jp/dp/login"
DROOM_ROOMLIST_URL = "https://anavi.daiwaliving.co.jp/dp/navi/room/RoomList/menu"
DROOM_TENPO        = os.getenv("DROOM_TENPO", "")
DROOM_TANTO        = os.getenv("DROOM_TANTO", "")
DROOM_PASS         = os.getenv("DROOM_PASS", "")

# ─── リアブロ (リアネットプロ) 認証情報 ─────────────────────
REABRO_LOGIN_URL = "https://www.realnetpro.com/index.php"
REABRO_BASE_URL  = "https://www.realnetpro.com"
REABRO_ID        = os.getenv("REABRO_ID", "")
REABRO_PASS      = os.getenv("REABRO_PASS", "")

# 都道府県コード（東建ルームサーチ用）
PREF_CODE = {
    "北海道": "01", "青森": "02", "岩手": "03", "宮城": "04",
    "秋田": "05", "山形": "06", "福島": "07", "茨城": "08",
    "栃木": "09", "群馬": "10", "埼玉": "11", "千葉": "12",
    "東京": "13", "神奈川": "14", "新潟": "15", "富山": "16",
    "石川": "17", "福井": "18", "山梨": "19", "長野": "20",
    "岐阜": "21", "静岡": "22", "愛知": "23", "三重": "24",
    "滋賀": "25", "京都": "26", "大阪": "27", "兵庫": "28",
    "奈良": "29", "和歌山": "30", "鳥取": "31", "島根": "32",
    "岡山": "33", "広島": "34", "山口": "35", "徳島": "36",
    "香川": "37", "愛媛": "38", "高知": "39", "福岡": "40",
    "佐賀": "41", "長崎": "42", "熊本": "43", "大分": "44",
    "宮崎": "45", "鹿児島": "46", "沖縄": "47",
}


# ══════════════════════════════════════════════════════════
#  ユーティリティ
# ══════════════════════════════════════════════════════════

def setup_dirs():
    INPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)
    SHOT_DIR.mkdir(parents=True, exist_ok=True)
    DETAIL_DIR.mkdir(parents=True, exist_ok=True)


def register_japanese_font() -> str:
    candidates = [
        # Linux (Docker/Railway) - IPA フォント（TTF形式、ReportLab対応）
        "/usr/share/fonts/opentype/ipafont-gothic/ipag.ttf",
        "/usr/share/fonts/opentype/ipafont-gothic/ipagp.ttf",
        "/usr/share/fonts/opentype/ipafont-mincho/ipam.ttf",
        "/usr/share/fonts/truetype/fonts-ipafont-gothic/ipag.ttf",
        "/usr/share/fonts/truetype/fonts-japanese-gothic.ttf",
        # Windows
        "C:/Windows/Fonts/meiryo.ttc",
        "C:/Windows/Fonts/msgothic.ttc",
        "C:/Windows/Fonts/YuGothM.ttc",
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("JaFont", path))
                print(f"フォント登録: {path}")
                return "JaFont"
            except Exception as e:
                print(f"フォント登録失敗 ({path}): {e}")
    print("日本語フォントが見つかりません。Helveticaを使用します")
    return "Helvetica"


def read_requests(excel_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        name = row[0]
        # ヒント行（「例：」で始まる行）と空行をスキップ
        if name is None or str(name).startswith("例：") or str(name).startswith("↑"):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


# ══════════════════════════════════════════════════════════
#  ATBB ログイン
# ══════════════════════════════════════════════════════════

async def login_atbb(page, shot_dir: Path) -> bool:
    print("── ATBBにログイン中 ──")
    await page.goto(ATBB_URL, wait_until="domcontentloaded", timeout=30000)
    await page.screenshot(path=str(shot_dir / "01_top.png"))

    # ログインフォームを探す
    id_selectors = [
        'input[name="userId"]', 'input[name="loginId"]', 'input[name="id"]',
        'input[name="memberId"]', 'input[id*="userId"]', 'input[id*="loginId"]',
        'input[type="text"]',
    ]
    pw_selectors = [
        'input[type="password"]',
        'input[name="password"]', 'input[name="pass"]',
    ]

    id_filled = False
    for sel in id_selectors:
        try:
            el = page.locator(sel).first
            if await el.is_visible(timeout=2000):
                await el.fill(ATBB_ID)
                print(f"  ID入力: {sel}")
                id_filled = True
                break
        except Exception:
            continue

    if not id_filled:
        print("  ⚠ ID入力フィールドが見つかりません")
        await page.screenshot(path=str(shot_dir / "01_error_no_id_field.png"))
        return False

    pw_filled = False
    for sel in pw_selectors:
        try:
            el = page.locator(sel).first
            if await el.is_visible(timeout=2000):
                await el.fill(ATBB_PASS)
                print(f"  PW入力: {sel}")
                pw_filled = True
                break
        except Exception:
            continue

    if not pw_filled:
        print("  ⚠ パスワード入力フィールドが見つかりません")
        return False

    await page.screenshot(path=str(shot_dir / "02_form_filled.png"))

    # ログインボタン
    submit_selectors = [
        'button[type="submit"]', 'input[type="submit"]',
        'button:has-text("ログイン")', 'a:has-text("ログイン")',
        'button:has-text("login")', 'button:has-text("Login")',
    ]
    submitted = False
    for sel in submit_selectors:
        try:
            el = page.locator(sel).first
            if await el.is_visible(timeout=2000):
                await el.click()
                submitted = True
                print(f"  ログインボタン: {sel}")
                break
        except Exception:
            continue

    if not submitted:
        await page.keyboard.press("Enter")
        print("  Enterキーで送信")

    try:
        await page.wait_for_load_state("networkidle", timeout=20000)
    except Exception:
        pass

    await page.screenshot(path=str(shot_dir / "03_after_login.png"))
    print(f"  ログイン後URL: {page.url}")

    # ログイン失敗チェック（URLが変わらない場合）
    if "login" in page.url.lower() or "auth" in page.url.lower():
        print("  ⚠ ログインに失敗した可能性があります")
        return False

    print("  ✓ ログイン成功")
    return True


# ══════════════════════════════════════════════════════════
#  ATBB 物件検索
# ══════════════════════════════════════════════════════════

async def search_atbb(page, ctx, area: str, rent_max: int, layout: str, shot_dir: Path) -> tuple[list[dict], object]:
    print(f"── ATBB 物件検索: area={repr(area)} / {rent_max}円 / {layout} ──")

    # エリアが空の場合は全国検索になるためスキップ
    if not area.strip():
        print("  ATBB: エリア未指定のため検索をスキップします")
        return [], page

    await page.screenshot(path=str(shot_dir / "04_logged_in.png"))

    properties = []

    try:
        # Step1: 「物件・会社検索」メニューをクリックしてサブメニューを開く
        try:
            el = page.locator('a:has-text("物件・会社検索")').first
            await el.wait_for(state="visible", timeout=5000)
            await el.click()
            print("  メニュークリック: 物件・会社検索")
        except Exception as e:
            print(f"  メニュークリック失敗: {e}")

        # クリック直後にメニューが開く（Riot.jsコンポーネント）
        await asyncio.sleep(1)
        await page.screenshot(path=str(shot_dir / "05_after_menu.png"))

        # Step2: 「流通物件検索」をクリック（新規タブが開く可能性あり）
        try:
            el = page.get_by_text("流通物件検索").first
            await el.wait_for(state="visible", timeout=3000)

            # 新規タブが開く場合に備えて待機
            async with ctx.expect_page(timeout=5000) as new_page_info:
                await el.click()
            new_page = await new_page_info.value
            await new_page.wait_for_load_state("domcontentloaded", timeout=15000)
            print(f"  新規タブで開きました: {new_page.url}")
            # 以降は新規タブで操作
            page = new_page
        except Exception as e:
            print(f"  新規タブ検出失敗 ({e}): 同一ページで継続")
            await asyncio.sleep(2)

        await asyncio.sleep(2)
        await page.screenshot(path=str(shot_dir / "05b_after_ryutsu_click.png"))
        print(f"  クリック後URL: {page.url}")

        # 「他のユーザーがATBBを利用中」ページへの対応
        if "ConcurrentLoginException" in page.url:
            print("  他セッション検出 → 強制ログイン処理")
            try:
                import re as _re
                force_info = await page.evaluate("""
                    () => {
                        const form = document.querySelector('form');
                        if (!form) return null;
                        const params = {};
                        Array.from(form.querySelectorAll('input')).forEach(i => {
                            if (i.name) params[i.name] = i.value;
                        });
                        // リンクも探す
                        const links = Array.from(document.querySelectorAll('a')).map(a => ({text:a.textContent.trim(), href:a.href}));
                        return {action: form.action, method: form.method, params: params, links: links.slice(0,5)};
                    }
                """)
                print(f"  フォーム情報: {force_info}")

                if force_info:
                    action = force_info['action']
                    # 現在URLのjsessionidを引き継ぐ
                    sid_match = _re.search(r'jsessionid=([A-Fa-f0-9]+)', page.url)
                    if sid_match and 'jsessionid' not in action:
                        action = action + f';jsessionid={sid_match.group(1)}'
                    # GETパラメータを追加
                    params = force_info.get('params') or {}
                    if params:
                        qs = '&'.join(f"{k}={v}" for k, v in params.items())
                        action = action + '?' + qs
                    print(f"  強制ログインURL: {action}")
                    await page.goto(action, wait_until="domcontentloaded", timeout=20000)
                else:
                    await page.evaluate("() => { const f=document.querySelector('form'); if(f) f.submit(); }")

                await asyncio.sleep(4)
                try:
                    await page.wait_for_load_state("domcontentloaded", timeout=15000)
                except Exception:
                    pass
                await asyncio.sleep(2)
                await page.screenshot(path=str(shot_dir / "05c_after_force_login.png"))
                print(f"  強制ログイン後URL: {page.url}")
            except Exception as e:
                print(f"  強制終了エラー: {e}")

        await asyncio.sleep(3)
        try:
            await page.wait_for_load_state("domcontentloaded", timeout=15000)
        except Exception:
            pass
        await asyncio.sleep(2)
        await page.screenshot(path=str(shot_dir / "06_ryutsu_search.png"))
        print(f"  流通物件検索URL: {page.url}")

        # ATBBの検索フォームを操作
        await asyncio.sleep(2)
        await page.screenshot(path=str(shot_dir / "06_ryutsu_search.png"))

        # Step1: ラジオボタンとラベルのマッピングを確認
        radio_label_map = await page.evaluate("""
            () => {
                const results = [];
                document.querySelectorAll('input[type=radio]').forEach(r => {
                    // 隣接するテキストを探す
                    let label = '';
                    const lb = document.querySelector('label[for="' + r.id + '"]');
                    if (lb) { label = lb.textContent.trim(); }
                    else {
                        let sib = r.nextSibling;
                        while (sib) {
                            if (sib.nodeType === 3) { label += sib.textContent.trim(); }
                            else if (sib.nodeType === 1) { label += sib.textContent.trim(); break; }
                            sib = sib.nextSibling;
                        }
                    }
                    results.push({name:r.name, value:r.value, label:label.substring(0,20)});
                });
                return results;
            }
        """)
        print(f"  ラジオ/ラベル対応: {radio_label_map[:10]}")

        # 「賃貸居住用」のvalue値を特定してクリック
        chintai_val = None
        for r in radio_label_map:
            if "賃貸居住用" in r.get("label","") or "賃貸居住" in r.get("label",""):
                chintai_val = r["value"]
                break
        if not chintai_val:
            # JSでラベルテキストから探す
            chintai_val = await page.evaluate("""
                () => {
                    const labels = Array.from(document.querySelectorAll('label,td,th'));
                    for (const el of labels) {
                        if (el.textContent.includes('賃貸居住用')) {
                            const input = el.querySelector('input[type=radio]');
                            if (input) return input.value;
                            // 前後のinputを探す
                            const prev = el.previousElementSibling;
                            if (prev && prev.tagName==='INPUT') return prev.value;
                        }
                    }
                    return null;
                }
            """)

        if chintai_val:
            await page.locator(f'input[name="atbbShumokuDaibunrui"][value="{chintai_val}"]').click()
            print(f"  賃貸居住用を選択: value={chintai_val}")
        else:
            print("  賃貸居住用ラジオが見つかりません。デフォルト値(06)を試みます")
            try:
                await page.locator('input[name="atbbShumokuDaibunrui"][value="06"]').click()
            except Exception:
                pass

        await page.screenshot(path=str(shot_dir / "07_chintai_selected.png"))
        await asyncio.sleep(1)

        # Step2: フリーワード検索フィールドに条件を入力して検索
        rent_man = rent_max // 10000
        free_word = f"{area} {rent_man}万円以下"
        freeword_sel = 'input[name*="freeword"], input[name*="freeWord"], input[placeholder*="万円"]'
        try:
            fw_el = page.locator(freeword_sel).first
            if await fw_el.is_visible(timeout=3000):
                await fw_el.fill(free_word)
                print(f"  フリーワード入力: {free_word}")
                # フリーワード横の検索ボタン
                fw_btn = page.locator('input[value="検索"]').last
                await fw_btn.click()
                print("  フリーワード検索実行")
            else:
                raise Exception("フリーワードフィールドが見つかりません")
        except Exception as e:
            print(f"  フリーワード検索失敗({e}): 所在地チェックボックスから検索")
            # 大阪府チェックボックスをクリック
            osaka_cb = await page.evaluate("""
                () => {
                    const cbs = Array.from(document.querySelectorAll('input[type=checkbox]'));
                    const osaka = cbs.find(c => {
                        const lb = document.querySelector('label[for="'+c.id+'"]');
                        return lb && lb.textContent.includes('大阪府');
                    }) || cbs.find(c => c.value && c.value.includes('27'));
                    if (osaka) { osaka.click(); return osaka.value; }
                    return null;
                }
            """)
            print(f"  大阪府チェックボックス: {osaka_cb}")
            # エリア検索ボタン
            for sel in ['input[type="submit"]', 'button:has-text("検索")']:
                try:
                    el = page.locator(sel).first
                    if await el.is_visible(timeout=2000):
                        await el.click()
                        break
                except Exception:
                    continue

        await asyncio.sleep(4)
        try:
            await page.wait_for_load_state("domcontentloaded", timeout=20000)
        except Exception:
            pass
        await asyncio.sleep(2)
        await page.screenshot(path=str(shot_dir / "08_search_results.png"))
        print(f"  検索結果URL: {page.url}")

        properties = await extract_properties(page, ctx, shot_dir)

    except Exception as e:
        print(f"  ⚠ 検索エラー: {e}")
        await page.screenshot(path=str(shot_dir / "search_error.png"))

    return properties, page


# ══════════════════════════════════════════════════════════
#  物件フィルタリング
# ══════════════════════════════════════════════════════════

def _parse_rent_yen(rent_str: str) -> int | None:
    """
    賃料文字列を円（int）に変換する。
    例: '7.5万円' → 75000, '75,000円' → 75000, '75000' → 75000
    解析不能の場合は None を返す。
    """
    if not rent_str:
        return None
    s = str(rent_str).strip()
    # "X.X万円" / "XX万円"
    m1 = re.search(r'(\d+\.?\d*)万円', s)
    if m1:
        return int(float(m1.group(1)) * 10000)
    # "XX,XXX円" / "XXXXX円"
    m2 = re.search(r'([\d,]+)円', s)
    if m2:
        return int(m2.group(1).replace(',', ''))
    # 数字のみ（5桁以上なら円とみなす）
    m3 = re.fullmatch(r'(\d+)', s)
    if m3:
        v = int(m3.group(1))
        return v if v >= 10000 else None
    return None


def _parse_age_years(age_str: str) -> int | None:
    """
    築年数文字列を年数（int）に変換する。
    '新築' → 0, '築5年' → 5, '築5年4ヶ月' → 5
    解析不能の場合は None を返す（フィルタをスキップ）。
    """
    if not age_str:
        return None
    s = str(age_str).strip()
    if '新築' in s:
        return 0
    m = re.search(r'築(\d+)年', s)
    if m:
        return int(m.group(1))
    return None


def filter_properties(
    props: list[dict],
    rent_max: int,
    max_age_years: int = 20,
) -> list[dict]:
    """
    物件リストから不適切な物件を除外する。

    除外条件:
    ① 家賃が低すぎる: 上限家賃の30%未満、かつ10,000円未満
       例: rent_max=100,000 → 30,000円未満の物件を除外
    ② 築年数が古すぎる: max_age_years 年超の物件を除外（デフォルト20年）
       ※ 築年数情報がない物件は除外しない
    """
    min_rent = max(10_000, int(rent_max * 0.30))
    filtered = []
    for p in props:
        name_short = p.get('name', '不明')[:25]
        rent_str = p.get('rent', '')
        rent_yen = _parse_rent_yen(rent_str)

        # ① 家賃チェック
        if rent_yen is not None and rent_yen < min_rent:
            min_man = min_rent // 10000
            print(f"  [フィルタ除外] {name_short}: 家賃{rent_str} → {rent_yen:,}円 < 下限{min_man}万円")
            continue

        # ② 築年数チェック
        age_str = p.get('age', '')
        age_years = _parse_age_years(age_str)
        if age_years is not None and age_years > max_age_years:
            print(f"  [フィルタ除外] {name_short}: {age_str} → {age_years}年 > 上限{max_age_years}年")
            continue

        filtered.append(p)

    removed = len(props) - len(filtered)
    if removed:
        print(f"  フィルタ結果: {len(props)}件 → {len(filtered)}件（{removed}件除外）")
    return filtered


def decode_price_code(code: str) -> str:
    """ATBBの価格コードをダウンロード&OCRでデコードして万円表示を返す"""
    import re as _re
    import urllib.request as _urlreq
    import io as _io
    from PIL import Image as _Image

    try:
        import ddddocr as _ocr_lib
    except ImportError:
        return ""

    try:
        url = f"https://d280xyghme9e5g.cloudfront.net/txt2img?f=2&v=2&m=mlb&d={code}"
        r = _urlreq.urlopen(url, timeout=8)
        data = r.read()

        img = _Image.open(_io.BytesIO(data))
        w, h = img.size
        bg = _Image.new("RGBA", (w, h), (255, 255, 255, 255))
        bg.paste(img, mask=img.split()[3])
        big = bg.convert("RGB").resize((w * 4, h * 4), _Image.LANCZOS)
        buf = _io.BytesIO()
        big.save(buf, "PNG")

        ocr = _ocr_lib.DdddOcr(show_ad=False)
        raw = ocr.classification(buf.getvalue())
        clean = _re.sub(r"[^\d.,]", "", raw).replace(",", ".")

        if "." in clean:
            price_num = clean  # already "X.XX" format
        elif len(clean) == 3 and clean.isdigit():
            price_num = f"{clean[0]}.{clean[1:]}"  # "380" → "3.80"
        elif len(clean) == 4 and clean.isdigit():
            price_num = f"{clean[0]}{clean[1]}.{clean[2:]}"  # "1200" → "12.00"
        else:
            price_num = clean

        return f"{price_num}万円" if price_num else ""
    except Exception:
        return ""


async def get_rent_from_detail(page, ctx, prop_index: int, shot_dir: Path) -> str:
    """詳細ボタンをクリックして賃料を取得する"""
    import re
    try:
        btn_id = f"shosai_{prop_index}"
        btn = page.locator(f"#{btn_id}")
        if not await btn.is_visible(timeout=3000):
            return ""

        # 新規タブが開く場合に備える
        try:
            async with ctx.expect_page(timeout=5000) as new_page_info:
                await btn.click()
            detail_page = await new_page_info.value
            await detail_page.wait_for_load_state("domcontentloaded", timeout=20000)
            await asyncio.sleep(3)
        except Exception:
            # 同一タブで遷移した場合
            await page.wait_for_load_state("domcontentloaded", timeout=15000)
            await asyncio.sleep(2)
            detail_page = page

        await detail_page.screenshot(path=str(shot_dir / f"detail_{prop_index}.png"))

        # 詳細ページから賃料を取得
        rent_text = await detail_page.evaluate("""
            () => {
                // 賃料ラベルの隣のセルまたは要素を探す
                const labels = Array.from(document.querySelectorAll('th,td,dt,label'));
                for (const el of labels) {
                    if (el.textContent.trim() === '賃料') {
                        // 隣の要素を取得
                        const next = el.nextElementSibling || el.parentElement?.nextElementSibling;
                        if (next) {
                            const t = next.textContent.replace(/\\s+/g,' ').trim();
                            if (t) return t;
                        }
                    }
                }
                // フォールバック: 万円を含む要素を探す
                const all = Array.from(document.querySelectorAll('*'));
                for (const el of all) {
                    if (el.children.length === 0) {
                        const t = el.textContent.trim();
                        if (/^\\d+(\\.\\d+)?万円$/.test(t)) return t;
                    }
                }
                // img alt textから試みる
                const imgs = Array.from(document.querySelectorAll('img[src*="txt2img"]'));
                for (const img of imgs) {
                    const src = img.src;
                    return '(画像)';
                }
                return '';
            }
        """)
        print(f"  詳細{prop_index}賃料: {rent_text}")

        # 新規タブの場合は閉じる
        if detail_page is not page:
            await detail_page.close()

        # 万円または円のパターンを抽出
        m = re.search(r'\d+\.?\d*万円|\d{3,6}円|\d{3,6}', rent_text)
        return m.group() if m else rent_text[:20]

    except Exception as e:
        print(f"  詳細{prop_index}取得エラー: {e}")
        return ""


async def get_floor_plan_image(page, ctx, prop_idx: int, shot_dir: Path) -> tuple[str, str]:
    """ATBBの詳細ボタンをクリック→同一タブ遷移→物件写真/間取り図取得→戻る。(path, label)を返す"""
    import urllib.request as _req
    out_path = str(shot_dir / f"madori_{prop_idx + 1}.png")
    photo_label = "物件写真"
    try:
        btn_exists = await page.evaluate(
            f"() => !!document.getElementById('shosai_{prop_idx}')"
        )
        if not btn_exists:
            print(f"  shosai_{prop_idx} が見つかりません（物件{prop_idx+1}）")
            return "", photo_label

        # クリック → 同一タブで詳細ページへ遷移
        await page.evaluate(f"document.getElementById('shosai_{prop_idx}').click()")
        try:
            await page.wait_for_load_state("domcontentloaded", timeout=20000)
        except Exception:
            pass
        await asyncio.sleep(3)
        print(f"  詳細URL 物件{prop_idx+1}: {page.url}")

        # ネットワークが落ち着くまで待つ（AJAXコンテンツのため）
        try:
            await page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass
        await asyncio.sleep(3)

        # 間取り図を優先取得（gazo-madori）→ 外観写真（gazo-gaikan）→ その他写真の順
        madori_result = await page.evaluate("""
            () => {
                function getImgUrl(sel) {
                    const li = document.querySelector(sel);
                    if (!li) return null;
                    const img = li.querySelector('img[src*="img4.athome.jp"]');
                    if (!img) return null;
                    // サムネイルURLの解像度を上げる
                    return img.src.replace(/[?&]height=\\d+/, '').replace(/[?&]width=\\d+/, '')
                               .replace(/[?&]margin=[^&]+/, '').replace(/[?&]dummy=[^&]+/, '')
                               .replace(/[?]+$/, '');
                }
                const madori = getImgUrl('li.gazo-madori');
                if (madori) return {url: madori, label: '間取り図'};
                const gaikan = getImgUrl('li.gazo-gaikan');
                if (gaikan) return {url: gaikan, label: '外観写真'};
                const naibu = getImgUrl('li.gazo-naibu');
                if (naibu) return {url: naibu, label: '室内写真'};
                // どのクラスでもない場合：最初のimg4.athome.jp画像
                const anyImg = document.querySelector('img[src*="img4.athome.jp"]');
                if (anyImg) {
                    const url = anyImg.src.replace(/[?&]height=\\d+/, '').replace(/[?&]width=\\d+/, '')
                                    .replace(/[?&]margin=[^&]+/, '').replace(/[?&]dummy=[^&]+/, '')
                                    .replace(/[?]+$/, '');
                    return {url: url, label: '物件写真'};
                }
                return null;
            }
        """)

        if madori_result and madori_result.get('url'):
            img_url = madori_result['url']
            photo_label = madori_result.get('label', '物件写真')
            try:
                _req.urlretrieve(img_url, out_path)
                print(f"  {photo_label}取得 物件{prop_idx+1}: ...{img_url[-50:]}")
            except Exception as dl_err:
                try:
                    fallback_url = img_url + '?height=400&width=400'
                    _req.urlretrieve(fallback_url, out_path)
                    print(f"  {photo_label}(サムネイル)取得 物件{prop_idx+1}")
                except Exception:
                    print(f"  画像ダウンロード失敗 物件{prop_idx+1}: {dl_err}")
                    out_path = ""
        else:
            print(f"  物件写真なし 物件{prop_idx+1}")
            out_path = ""

        # 検索結果ページに戻る
        await page.go_back()
        try:
            await page.wait_for_load_state("domcontentloaded", timeout=15000)
        except Exception:
            pass
        await asyncio.sleep(2)

    except Exception as e:
        print(f"  写真取得エラー 物件{prop_idx+1}: {e}")
        try:
            await page.go_back()
            await asyncio.sleep(2)
        except Exception:
            pass
        out_path = ""

    return out_path, photo_label


# ══════════════════════════════════════════════════════════
#  東建ルームサーチ ログイン・検索・物件抽出
# ══════════════════════════════════════════════════════════

def detect_prefecture(area_text: str) -> str:
    """エリア文字列から東建ルームサーチの都道府県コードを特定する"""
    for name, code in PREF_CODE.items():
        if name in area_text:
            return code
    # 主要市区から推定
    if "大阪市" in area_text:
        return "27"
    if "神戸市" in area_text:
        return "28"
    if any(w in area_text for w in ["新宿区", "渋谷区", "品川区", "千代田区", "中央区",
                                     "港区", "文京区", "台東区", "墨田区", "江東区",
                                     "目黒区", "大田区", "世田谷区", "中野区", "杉並区",
                                     "豊島区", "荒川区", "板橋区", "練馬区", "足立区"]):
        return "13"
    return ""


async def download_atbb_print_pdf(page, ctx, playwright_instance,
                                  prop_idx: int, prop: dict,
                                  out_dir: Path, font_name: str) -> str:
    """ATBBインフォシートのPDF出力ボタンをクリックして社宅.com版PDF（手数料非表示）を取得する。
    Chromeの組み込みPDFビューア(chrome-extension)経由でサーブされるPDFバイトを
    ctx.on('response') でインターセプトして保存する。"""
    prop_safe = re.sub(r'[\\/:*?"<>|]', '_', prop.get('name', f'物件{prop_idx+1}'))[:50]
    out_pdf   = str(out_dir / f"{prop_safe}_物件詳細.pdf")

    info_tab = None
    new_pg   = None
    _on_response = None

    try:
        # ─── ① shosai_N クリック → 詳細ページへ ───
        btn_exists = await page.evaluate(
            f"() => !!document.getElementById('shosai_{prop_idx}')")
        if not btn_exists:
            print(f"  shosai_{prop_idx} が見つかりません")
            return ""

        await page.evaluate(f"document.getElementById('shosai_{prop_idx}').click()")
        await page.wait_for_load_state("domcontentloaded", timeout=20000)
        await asyncio.sleep(3)
        try:
            await page.wait_for_load_state("networkidle", timeout=8000)
        except Exception:
            pass
        await asyncio.sleep(2)
        print(f"  ATBB詳細 物件{prop_idx+1}: {page.url[:70]}")

        # ─── ② インフォシートボタン → 新規タブ（React SPA）を開く ───
        btn = page.locator('#infoSheetButtonTop_0')
        if not await btn.is_visible(timeout=3000):
            btn = page.locator('button:has-text("インフォシート")').first
        if not await btn.is_visible(timeout=3000):
            print(f"  ⚠ インフォシートボタンが見つかりません 物件{prop_idx+1}")
            await page.go_back()
            await asyncio.sleep(2)
            return ""

        try:
            async with ctx.expect_page(timeout=10000) as np_info:
                await btn.click()
            info_tab = await np_info.value
            await info_tab.wait_for_load_state("domcontentloaded", timeout=15000)
        except Exception as tab_err:
            print(f"  新規タブ検出失敗({tab_err}) → ctx.pages を探索")
            for p in ctx.pages:
                if "infosheets" in p.url or "zmn.atbb" in p.url:
                    info_tab = p
                    break

        if not info_tab:
            print(f"  ⚠ インフォシートタブが取得できませんでした")
            await page.go_back()
            await asyncio.sleep(2)
            return ""

        # React SPA の描画を待つ
        await asyncio.sleep(6)
        print(f"  インフォシートURL: {info_tab.url[:80]}")

        # ─── ③ PDFレスポンスをインターセプトするリスナーを登録 ───
        pdf_captured: list = []

        async def _on_response(response):
            skip_exts = ('.js', '.css', '.png', '.jpg', '.ico',
                         '.woff', '.woff2', '.svg', '.gif')
            url = response.url
            if any(url.endswith(e) for e in skip_exts):
                return
            try:
                body = await response.body()
                if body[:4] == b'%PDF':
                    pdf_captured.append(body)
                    print(f"  ★ PDFキャプチャ: {len(body):,}B  {url[:70]}")
            except Exception:
                pass

        ctx.on('response', _on_response)

        # ─── ④ infosheets JSON の feeDispFlg を false に書き換え ───
        try:
            await info_tab.evaluate("""
                () => {
                    const inp = document.querySelector('input[name="infosheets"]');
                    if (!inp) return;
                    try {
                        const arr = JSON.parse(inp.value);
                        arr.forEach(item => { item.feeDispFlg = false; });
                        inp.value = JSON.stringify(arr);
                    } catch(e) {}
                }
            """)
            print(f"  feeDispFlg → false に書き換え完了")
        except Exception as fe:
            print(f"  feeDispFlg書き換え失敗: {fe}")

        # ─── ⑤ PDF出力ボタンをクリック → 新規タブ（Chrome PDFビューア）が開く ───
        try:
            async with ctx.expect_page(timeout=15000) as pdf_pg_info:
                await info_tab.click('#button-pdf-format')
            new_pg = await pdf_pg_info.value
            # Chrome PDFビューア拡張がPDFをサーブするまで待つ
            await asyncio.sleep(10)
            print(f"  PDFタブURL: {new_pg.url[:70]}")
        except Exception as pdf_tab_err:
            print(f"  PDFタブ待機エラー: {pdf_tab_err}")
            await asyncio.sleep(10)  # タイムアウトしてもレスポンスは来ている場合がある

        # ─── ⑥ キャプチャしたPDFを保存 ───
        saved = False
        if pdf_captured:
            with open(out_pdf, 'wb') as f:
                f.write(pdf_captured[0])
            print(f"  ✓ ATBBインフォシートPDF保存(社宅.com版): {out_pdf} ({len(pdf_captured[0]):,}B)")
            saved = True
        else:
            # フォールバック: info_tab を page.pdf() で印刷
            print(f"  PDFキャプチャ失敗 → info_tab.pdf() フォールバック")
            try:
                await info_tab.pdf(path=out_pdf, format="A4", print_background=True)
                print(f"  ✓ ATBBインフォシートPDF保存(フォールバック): {out_pdf}")
                saved = True
            except Exception as fb_err:
                print(f"  フォールバックPDF失敗: {fb_err}")

        # ─── ⑦ クリーンアップ ───
        try:
            ctx.remove_listener('response', _on_response)
        except Exception:
            pass
        try:
            if new_pg and not new_pg.is_closed():
                await new_pg.close()
        except Exception:
            pass
        try:
            if info_tab and not info_tab.is_closed():
                await info_tab.close()
        except Exception:
            pass

        # 詳細ページ → 検索結果に戻る
        try:
            await page.go_back()
            await page.wait_for_load_state("domcontentloaded", timeout=15000)
            await asyncio.sleep(2)
        except Exception:
            pass

        print(f"  検索結果に戻った: {page.url[:60]}")
        return out_pdf if saved else ""

    except Exception as e:
        print(f"  ⚠ ATBB PDF取得エラー 物件{prop_idx+1}: {e}")
        import traceback; traceback.print_exc()
        # クリーンアップ
        if _on_response is not None:
            try:
                ctx.remove_listener('response', _on_response)
            except Exception:
                pass
        try:
            if new_pg and not new_pg.is_closed():
                await new_pg.close()
        except Exception:
            pass
        try:
            if info_tab and not info_tab.is_closed():
                await info_tab.close()
        except Exception:
            pass
        try:
            await page.go_back()
            await asyncio.sleep(2)
        except Exception:
            pass
        return ""

async def download_homemate_detail_pdf(page, playwright, detail_href: str, out_dir: Path) -> str:
    """東建ルームサーチ詳細ページからPDF全印刷URLを取得してPDFを保存する"""
    if not detail_href:
        return ""
    try:
        await page.goto(detail_href, wait_until="domcontentloaded", timeout=20000)
        await asyncio.sleep(3)

        # "PDF全印刷" のリンクURLを抽出（pdfset.asp?bn=...&id=...）
        pdf_url = await page.evaluate("""
            () => {
                for (const a of document.querySelectorAll('a')) {
                    const t = a.textContent.trim();
                    const h = a.href || '';
                    if (t.includes('PDF全印刷') && h.includes('pdfset.asp')) return h;
                }
                // 全印刷がなければ最初の pdfset.asp リンク
                for (const a of document.querySelectorAll('a')) {
                    if ((a.href || '').includes('pdfset.asp')) return a.href;
                }
                return null;
            }
        """)

        if not pdf_url:
            print(f"  ⚠ PDFリンクが見つかりません: {detail_href}")
            return ""

        bn_m = re.search(r'bn=(\d+)', pdf_url)
        bn = bn_m.group(1) if bn_m else "bukken"
        out_path = str(out_dir / f"物件詳細_{bn}.pdf")
        print(f"  PDF URL取得: ...{pdf_url[-55:]}")

        # ── urllib で直接ダウンロード（ID トークンが URL に含まれるため Cookie 不要）──
        import urllib.request as _urlreq
        downloaded = False
        try:
            req = _urlreq.Request(pdf_url, headers={
                'User-Agent': ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                               'AppleWebKit/537.36 (KHTML, like Gecko) '
                               'Chrome/120.0.0.0 Safari/537.36'),
                'Referer': detail_href,
                'Accept': 'application/pdf,*/*;q=0.9',
            })
            with _urlreq.urlopen(req, timeout=30) as resp:
                body = resp.read()
            if body[:4] == b'%PDF':
                with open(out_path, 'wb') as f:
                    f.write(body)
                print(f"  ✓ 物件詳細PDF保存(直接DL): {out_path}")
                downloaded = True
            else:
                print(f"  ⚠ PDF形式ではありません（先頭: {body[:30]}）→ 別手段へ")
        except Exception as dl_err:
            print(f"  urllib DLエラー: {dl_err}")

        if not downloaded:
            # フォールバック：ヘッドレスブラウザでPDF印刷
            try:
                cookies = await page.context.cookies()
                headless_browser = await playwright.chromium.launch(headless=True)
                headless_ctx = await headless_browser.new_context(
                    viewport={"width": 1280, "height": 900}, locale="ja-JP")
                await headless_ctx.add_cookies(cookies)
                headless_page = await headless_ctx.new_page()
                await headless_page.goto(pdf_url, wait_until="domcontentloaded", timeout=30000)
                await asyncio.sleep(2)
                await headless_page.pdf(path=out_path, format="A4", print_background=True)
                await headless_browser.close()
                print(f"  ✓ 物件詳細PDF保存(印刷): {out_path}")
                downloaded = True
            except Exception as pdf_err:
                print(f"  PDF印刷も失敗({pdf_err}): スクリーンショットで代替")
                ss_path = out_path.replace('.pdf', '_fullpage.png')
                await page.screenshot(path=ss_path, full_page=True)
                out_path = ss_path
                print(f"  ✓ スクリーンショット保存: {out_path}")

        return out_path

    except Exception as e:
        print(f"  ⚠ 物件詳細PDF取得エラー: {e}")
        import traceback; traceback.print_exc()
        return ""


def merge_pdf_with_map_image(detail_pdf_path: str, map_img_path: str,
                              out_path: str, font_name: str,
                              commute_method: str = '',
                              workplace: str = '') -> str:
    """物件詳細PDFの末尾に通勤ルート地図ページを追加して地図付きPDFを生成する"""
    import io as _io
    try:
        W, H = A4
        M = 15 * mm

        # 地図ページのヘッダー文字列を組み立てる
        header_parts = ["通勤ルート地図"]
        if commute_method:
            header_parts.append(f"（{commute_method}）")
        if workplace:
            short_wp = workplace[:30] + ('…' if len(workplace) > 30 else '')
            header_parts.append(f"  勤務地: {short_wp}")
        header_text = ''.join(header_parts)

        # 地図画像を1ページのPDFに変換
        map_buf = _io.BytesIO()
        c = canvas.Canvas(map_buf, pagesize=A4)
        c.setFont(font_name, 11)
        c.setFillColor(colors.HexColor("#37474f"))
        c.drawString(M, H - M, header_text)

        img = Image.open(map_img_path)
        iw, ih = img.size
        max_w = W - 2 * M
        max_h = H - 2 * M - 15 * mm
        scale = min(max_w / iw, max_h / ih)
        dw, dh = iw * scale, ih * scale
        c.drawImage(map_img_path, M, H - M - 15 * mm - dh, width=dw, height=dh)
        c.save()
        map_buf.seek(0)

        # 物件詳細PDF + 地図ページ をマージ
        writer = PdfWriter()
        if detail_pdf_path.endswith('.pdf') and os.path.exists(detail_pdf_path):
            reader = PdfReader(detail_pdf_path)
            for p in reader.pages:
                writer.add_page(p)

        map_reader = PdfReader(map_buf)
        for p in map_reader.pages:
            writer.add_page(p)

        with open(out_path, 'wb') as f:
            writer.write(f)

        print(f"  ✓ 地図付きPDF保存: {out_path}")
        return out_path

    except Exception as e:
        print(f"  ⚠ PDF合成エラー: {e}")
        import traceback; traceback.print_exc()
        return detail_pdf_path


async def login_homemate(page, shot_dir: Path) -> bool:
    """東建ルームサーチにログインしてtop.aspへ遷移する"""
    print("── 東建ルームサーチにログイン中 ──")
    await page.goto(HM_URL, wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(2)
    await page.screenshot(path=str(shot_dir / "hm_01_top.png"))

    try:
        await page.locator('input[name="id"]').fill(HM_ID)
        await page.locator('input[name="pw"]').fill(HM_PASS)
        await page.locator('#btn_login a').click()
        print("  ログインボタンクリック")
    except Exception as e:
        print(f"  ⚠ ログインフォーム操作エラー: {e}")
        return False

    try:
        await page.wait_for_load_state("domcontentloaded", timeout=15000)
    except Exception:
        pass
    await asyncio.sleep(3)
    await page.screenshot(path=str(shot_dir / "hm_02_after_login.png"))

    if "top.asp" in page.url:
        print(f"  ✓ ログイン成功: {page.url}")
        return True
    else:
        print(f"  ⚠ ログイン失敗: {page.url}")
        return False


async def search_homemate(page, area: str, shot_dir: Path,
                          work_address: str = '') -> bool:
    """東建ルームサーチで都道府県→市→区の順に選択して検索実行"""
    import re as _re
    print(f"── 東建 エリア検索: {area} ──")

    # top.aspに居なければ遷移
    if "top.asp" not in page.url:
        await page.goto(HM_URL + "top.asp", wait_until="domcontentloaded", timeout=20000)
        await asyncio.sleep(2)

    # 都道府県コード特定（city名だけでは不明な場合は work_address で補う）
    pref_code = detect_prefecture(area)
    if not pref_code and work_address:
        pref_code = detect_prefecture(work_address)
    if not pref_code:
        print(f"  ⚠ 都道府県を特定できません: {area}")
        return False

    # ① 都道府県選択 → getCitySb() が走り市リストがAJAXロードされる
    await page.select_option('select[name="prf"]', pref_code)
    await asyncio.sleep(3)

    # ② 市（citysb）選択
    citysb_options = await page.evaluate("""
        () => Array.from(document.querySelector('select[name="citysb"]').options)
            .map(o => ({v: o.value, t: o.text.trim()}))
    """)
    print(f"  市グループ: {len(citysb_options)}件")

    area_norm = _re.sub(r'[　\s]', '', area)
    target_city_val = None
    for o in citysb_options:
        if o['t'] and o['t'] in area_norm:
            target_city_val = o['v']
            print(f"  市選択: {o['t']} (value={o['v']})")
            break

    if not target_city_val and len(citysb_options) > 1:
        # 先頭（空の選択肢を除く最初の実選択肢）
        for o in citysb_options:
            if o['v']:
                target_city_val = o['v']
                print(f"  市選択(先頭): {o['t']}")
                break

    if target_city_val:
        await page.select_option('select[name="citysb"]', target_city_val)
        await asyncio.sleep(3)  # getCity() が走る

        # ③ 区・町（seljiscd）選択
        seljiscd_options = await page.evaluate("""
            () => Array.from(document.querySelector('select[name="seljiscd"]').options)
                .map(o => ({v: o.value, t: o.text.trim()}))
        """)
        print(f"  区・町: {len(seljiscd_options)}件")

        target_ward_val = None
        for o in seljiscd_options:
            if o['t'] and o['t'] in area_norm:
                target_ward_val = o['v']
                print(f"  区選択: {o['t']} (value={o['v']})")
                break

        if not target_ward_val and seljiscd_options:
            for o in seljiscd_options:
                if o['v']:
                    target_ward_val = o['v']
                    print(f"  区選択(先頭): {o['t']}")
                    break

        if target_ward_val:
            await page.select_option('select[name="seljiscd"]', target_ward_val)
            await asyncio.sleep(1)
            # addlist() で stjiscd に追加
            await page.evaluate("""
                () => {
                    const f = document.frmCond;
                    if (typeof addlist === 'function') {
                        addlist(f.seljiscd, f.stjiscd, 'jiscd');
                    } else {
                        const sel = f.seljiscd;
                        const stock = f.stjiscd;
                        const opt = sel.options[sel.selectedIndex];
                        if (opt) stock.add(new Option(opt.text, opt.value));
                    }
                }
            """)
            await asyncio.sleep(1)

    await page.screenshot(path=str(shot_dir / "hm_03_form_ready.png"))

    # ④ 検索実行
    await page.evaluate("""
        () => {
            const stock = document.frmCond.stjiscd;
            for (let i = 0; i < stock.options.length; i++) {
                stock.options[i].selected = true;
            }
            const ok = (typeof nextSubmit === 'function') ? nextSubmit() : true;
            if (ok !== false) document.frmCond.submit();
        }
    """)

    try:
        await page.wait_for_load_state("domcontentloaded", timeout=20000)
    except Exception:
        pass
    await asyncio.sleep(4)
    await page.screenshot(path=str(shot_dir / "hm_04_result.png"))
    print(f"  検索結果URL: {page.url}")
    return True


async def extract_homemate_properties(page, ctx, shot_dir: Path) -> list[dict]:
    """東建ルームサーチの検索結果ページ(#ta_bukkenlist)から物件情報を抽出する"""
    import urllib.request as _req
    properties = []
    print("  東建物件情報を抽出中...")

    try:
        # #ta_bukkenlist テーブルから構造化データを取得するJS
        rows_data = await page.evaluate("""
            () => {
                // 号室・階数パーサー
                // セルテキスト "100210/10階角部屋" → {room:"1002", floor:"10/10階"}
                function parseRoomFloor(txt) {
                    const slash = txt.indexOf('/');
                    if (slash < 0) return {room: txt.replace(/\\D/g,'').slice(0,4), floor: ''};
                    const kai = txt.indexOf('階', slash);
                    if (kai < 0) return {room: txt.replace(/\\D/g,'').slice(0,4), floor: ''};
                    const total = parseInt(txt.slice(slash+1, kai));
                    if (isNaN(total)) return {room: txt, floor: ''};
                    // current floor: 最大2桁 ≤ total を "/" の直前から逆算
                    let curr = NaN, re = slash;
                    if (slash >= 2) {
                        const t2 = parseInt(txt.slice(slash-2, slash));
                        if (!isNaN(t2) && t2 > 0 && t2 <= total) { curr = t2; re = slash-2; }
                    }
                    if (isNaN(curr) && slash >= 1) {
                        const t1 = parseInt(txt.slice(slash-1, slash));
                        if (!isNaN(t1) && t1 > 0 && t1 <= total) { curr = t1; re = slash-1; }
                    }
                    return {room: txt.slice(0, re).trim(), floor: isNaN(curr)?'':(curr+'/'+total+'階')};
                }

                const table = document.getElementById('ta_bukkenlist');
                if (!table) return [];

                const props = [];
                let cur_name='', cur_addr='', cur_station='', cur_img='';
                const rows = Array.from(table.rows);

                for (let i = 1; i < rows.length; i++) {
                    const cells = Array.from(rows[i].cells);
                    const is_top = cells.length >= 10 ||
                                   (cells[0] && cells[0].className.includes('td_top'));

                    let room='', floor='', rent='', layout='', area='', built='', href='';

                    if (is_top) {
                        // ── 建物ヘッダー行 (11セル) ──
                        const img_el = cells[0] ? cells[0].querySelector('img') : null;
                        if (img_el && img_el.src) cur_img = img_el.src;

                        // cell1: 建物名+住所+最寄駅
                        const c1 = cells[1];
                        if (c1) {
                            const full = c1.textContent.replace(/\\s+/g,' ').trim();
                            const lnk  = c1.querySelector('a');
                            const lnk_txt = lnk ? lnk.textContent.replace(/\\s+/g,' ').trim() : '';
                            if (lnk_txt && full.includes(lnk_txt)) {
                                cur_name = full.slice(0, full.indexOf(lnk_txt)).trim();
                                const rest = full.slice(full.indexOf(lnk_txt));
                                const sm = rest.match(/(OsakaMetro|JR|近鉄|阪急|阪神|地下鉄|南海|京阪|東急|西鉄|メトロ|都営|東京メトロ).+?徒歩\\d+分/);
                                if (sm) {
                                    cur_addr = rest.slice(0, rest.indexOf(sm[0])).trim();
                                    cur_station = sm[0].replace(/\\s+/g,' ').trim();
                                } else { cur_addr = rest.trim(); cur_station = ''; }
                            } else { cur_name = full; }
                        }

                        // cell2: 号室+階建 (例: "100210/10階角部屋")
                        const c2t = cells[2] ? cells[2].textContent.trim() : '';
                        const rf2 = parseRoomFloor(c2t);
                        room = rf2.room; floor = rf2.floor;

                        // cell3: 賃料
                        const c3t = cells[3] ? cells[3].textContent.trim() : '';
                        const rm3 = c3t.match(/(\\d+\\.?\\d*)万円/);
                        rent = rm3 ? rm3[1]+'万円' : '';

                        // cell7: 間取り+面積
                        const c7t = cells[7] ? cells[7].textContent.trim() : '';
                        const lm7 = c7t.match(/(\\d[LKDS]+K?|ワンルーム|1R)/i);
                        const am7 = c7t.match(/(\\d+\\.?\\d*)m²/);
                        layout = lm7 ? lm7[1] : '';
                        area   = am7 ? am7[1]+'㎡' : '';

                        // cell8: "2021/052026/07/30" → 完成年月
                        const c8t = cells[8] ? cells[8].textContent.trim() : '';
                        const bm8 = c8t.match(/^(\\d{4})\\/(\\d{2})/);
                        built = bm8 ? bm8[1]+'/'+bm8[2] : '';

                        // cell10: 詳細リンク
                        const dl10 = cells[10] ? cells[10].querySelector('a') : null;
                        href = dl10 ? dl10.href : '';

                    } else {
                        // ── 同建物サブ行 (9セル) ──
                        // cell0: 号室+階建 (例: "3023/10階角部屋")
                        const c0t = cells[0] ? cells[0].textContent.trim() : '';
                        const rf0 = parseRoomFloor(c0t);
                        room = rf0.room; floor = rf0.floor;

                        // cell1: 賃料
                        const c1t = cells[1] ? cells[1].textContent.trim() : '';
                        const rm1 = c1t.match(/(\\d+\\.?\\d*)万円/);
                        rent = rm1 ? rm1[1]+'万円' : '';

                        // cell5: 間取り+面積
                        const c5t = cells[5] ? cells[5].textContent.trim() : '';
                        const lm5 = c5t.match(/(\\d[LKDS]+K?|ワンルーム|1R)/i);
                        const am5 = c5t.match(/(\\d+\\.?\\d*)m²/);
                        layout = lm5 ? lm5[1] : '';
                        area   = am5 ? am5[1]+'㎡' : '';

                        // cell6: 完成年月
                        const c6t = cells[6] ? cells[6].textContent.trim() : '';
                        const bm6 = c6t.match(/^(\\d{4})\\/(\\d{2})/);
                        built = bm6 ? bm6[1]+'/'+bm6[2] : '';

                        // cell8: 詳細リンク
                        const dl8 = cells[8] ? cells[8].querySelector('a') : null;
                        href = dl8 ? dl8.href : '';
                    }

                    // 築年数計算
                    let age = '';
                    if (built) {
                        const yr = parseInt(built.split('/')[0]);
                        if (!isNaN(yr)) age = (2026-yr) <= 0 ? '新築' : '築'+(2026-yr)+'年';
                    }

                    props.push({
                        name: cur_name, address: cur_addr, station: cur_station,
                        room: room, floor: floor, rent: rent,
                        layout: layout, area: area, age: age,
                        detail_href: href, img_url: cur_img
                    });
                }
                return props;
            }
        """)

        print(f"  東建テーブル解析: {len(rows_data)}件")
        for i, p in enumerate(rows_data, 1):
            print(f"  物件{i}: {p.get('name','')} {p.get('room','')}号室 / {p.get('rent','')} / {p.get('address','')[:25]}")

        for i, r in enumerate(rows_data[:3]):
            room_label = f" {r['room']}号室" if r.get('room') else ""
            floor_label = f" ({r['floor']})" if r.get('floor') else ""
            prop = {
                "name": f"{r['name']}{room_label}{floor_label}".strip(),
                "address": r.get('address', ''),
                "rent": r.get('rent', ''),
                "layout": r.get('layout', ''),
                "area": r.get('area', ''),
                "age": r.get('age', ''),
                "station": r.get('station', ''),
                "madori_path": "",
                "photo_label": "外観写真",
                "detail_href": r.get('detail_href', ''),
                "source": "homemate",
                "hm_idx": i,
            }

            # 外観写真を取得（結果ページの pic1.homemate.co.jp 画像 or 詳細ページ）
            img_url = r.get('img_url', '')
            if not img_url and r.get('detail_href'):
                # 詳細ページから写真を取得
                try:
                    await page.goto(r['detail_href'], wait_until="domcontentloaded", timeout=20000)
                    await asyncio.sleep(3)
                    img_url = await page.evaluate("""
                        () => {
                            // pic1.homemate.co.jp の画像を優先
                            const imgs = Array.from(document.querySelectorAll('img'))
                                .filter(img => img.src.includes('homemate.co.jp') ||
                                               img.src.includes('pic1.') ||
                                               img.src.includes('/ice/'))
                                .sort((a,b) => b.naturalWidth - a.naturalWidth);
                            if (imgs.length) return imgs[0].src;
                            // 間取り図
                            const madori = document.querySelector('img[alt*="間取"], img[alt*="Floor"]');
                            if (madori) return madori.src;
                            return '';
                        }
                    """)
                    await page.go_back()
                    try:
                        await page.wait_for_load_state("domcontentloaded", timeout=15000)
                    except Exception:
                        pass
                    await asyncio.sleep(2)
                except Exception as ph_err:
                    print(f"  詳細ページ取得エラー 物件{i+1}: {ph_err}")

            if img_url:
                out_path = str(shot_dir / f"hm_photo_{i+1}.jpg")
                try:
                    _req.urlretrieve(img_url, out_path)
                    prop['madori_path'] = out_path
                    print(f"  写真取得(東建) 物件{i+1}: {img_url[-50:]}")
                except Exception as dl_err:
                    print(f"  写真DLエラー 物件{i+1}: {dl_err}")

            properties.append(prop)

    except Exception as e:
        print(f"  ⚠ 東建抽出エラー: {e}")
        import traceback
        traceback.print_exc()

    return properties


async def extract_properties(page, ctx, shot_dir: Path) -> list[dict]:
    """ATBB物件一覧ページから情報を抽出する"""
    properties = []
    print("  物件情報を抽出中...")

    try:
        # ページ全文を取得してテキストパース
        full_page_text = await page.evaluate("() => document.body.innerText")
        properties = parse_atbb_full_text(full_page_text)

        # 賃料コード（kakakuChinryoImage の引数）を取得してOCRでデコード
        price_codes = await page.evaluate("""
            () => {
                const data = {};
                document.querySelectorAll('script:not([src])').forEach(s => {
                    const m = s.textContent.match(/kakakuChinryoImage\\([^,]+,\\s*"(\\d+)",\\s*"([^"]+)"\\)/);
                    if (m && !(m[1] in data)) data[m[1]] = m[2];
                });
                return data;
            }
        """)
        print(f"  賃料コード: {list(price_codes.items())[:3]}")

        for i, prop in enumerate(properties):
            code = price_codes.get(str(i))
            if code:
                rent = decode_price_code(code)
                if rent:
                    prop['rent'] = rent
                    print(f"  物件{i+1}賃料: {rent}")
            if not prop.get('rent'):
                prop['rent'] = "要確認"

        # 物件写真/間取り図を各詳細ページから取得
        print("  物件写真を取得中...")
        for i, prop in enumerate(properties):
            madori_path, photo_label = await get_floor_plan_image(page, ctx, i, shot_dir)
            prop['madori_path'] = madori_path
            prop['photo_label'] = photo_label

    except Exception as e:
        print(f"  ⚠ 抽出エラー: {e}")
        import traceback
        traceback.print_exc()

    return properties


def parse_atbb_full_text(text: str) -> list[dict]:
    """ATBBページのテキスト全体から物件情報をパースする"""
    import re
    properties = []
    # No.X から始まる物件ブロックを分割 (No.1, No.2, No.10 etc)
    blocks = re.split(r'(?=\bNo\.\d+\b)', text)
    blocks = [b for b in blocks if re.match(r'\bNo\.\d+\b', b.strip())][:3]
    print(f"  物件ブロック数: {len(blocks)}")

    for idx, block in enumerate(blocks, 1):
        lines = [l.strip() for l in block.split('\n') if l.strip()]
        prop = {
            "name": "",
            "address": "",
            "rent": "",
            "layout": "",
            "area": "",
            "age": "",
            "station": "",
        }
        # 物件種目パターン（スキップ対象）
        type_pat = re.compile(r'^(貸|売|賃貸|分譲|新築)(マンション|戸建|アパート|一軒家|事業用|土地|駐車場)')

        for i, line in enumerate(lines[:50]):
            # No.X 行の後に物件種目行をスキップして物件名を探す
            if re.match(r'\bNo\.\d+\b', line) and not prop["name"]:
                for j in range(i+1, min(i+6, len(lines))):
                    candidate = lines[j]
                    # 種目行・日付行・フラグ行はスキップ
                    if type_pat.match(candidate):
                        continue
                    if re.match(r'^(公開日|画像|新着|No\.)', candidate):
                        continue
                    if 3 < len(candidate) < 70:
                        # 読み仮名（カタカナ）の括弧部分を除去
                        prop["name"] = re.sub(r'[（(][ァ-ンヴー\s]+[）)]', '', candidate).strip()
                        break

            # 家賃: 万円表記 or 数万円レンジ (3万〜20万が賃貸の現実的な範囲)
            if not prop["rent"]:
                m = re.search(r'(\d+\.?\d*)\s*万円', line)
                if m:
                    val = float(m.group(1))
                    if 2 <= val <= 30:  # 2万〜30万の範囲のみ家賃とみなす
                        prop["rent"] = f"{m.group(1)}万円"
                # 円単位の場合 (20000〜200000)
                if not prop["rent"]:
                    m2 = re.search(r'([2-9]\d{4}|1\d{5}|20\d{4})\s*円', line.replace(',', ''))
                    if m2:
                        prop["rent"] = f"{int(m2.group(1)):,}円"

            # 住所
            if not prop["address"] and re.search(r'[都道府県]|市[区町村]|丁目|\d+[-－]\d+', line):
                if not re.search(r'万円|㎡|築|徒歩', line):  # 家賃・面積行を除外
                    prop["address"] = re.sub(r'\s*(地図|MAP|map)\s*$', '', line).strip()

            # 間取り
            if not prop["layout"]:
                m = re.search(r'(\d[LKDS]+K?|ワンルーム)', line, re.IGNORECASE)
                if m and len(m.group()) <= 6:
                    prop["layout"] = m.group()

            # 面積
            if not prop["area"]:
                m = re.search(r'(\d+\.?\d*)\s*㎡', line)
                if m:
                    prop["area"] = f"{m.group(1)}㎡"

            # 築年数（"築年月 1986/07" 形式 or "新築/築XX年" 形式）
            if not prop["age"]:
                m = re.search(r'築年月\s*(\d{4})/(\d{2})', line)
                if m:
                    built_year = int(m.group(1))
                    age_years = 2026 - built_year
                    prop["age"] = f"築{age_years}年" if age_years > 0 else "新築"
                else:
                    m2 = re.search(r'(新築|築\d+年?)', line)
                    if m2:
                        prop["age"] = m2.group()

            # 最寄駅（ATBB形式: "交通\tＪＲ大阪環状線 福島 徒歩11分"）
            if not prop["station"] and re.search(r'徒歩\d+分|バス\d+分', line):
                station_text = re.sub(r'^交通\s*', '', line).strip()
                prop["station"] = station_text[:50]

        if not prop["name"] and lines:
            prop["name"] = lines[1] if len(lines) > 1 else lines[0]

        properties.append(prop)
        print(f"  物件{idx}: {prop['name'][:30]} / {prop['rent']} / {prop['address'][:30]}")

    return properties


# ══════════════════════════════════════════════════════════
#  D-Room (大和リビング) ログイン・検索・Droomシート一括出力
# ══════════════════════════════════════════════════════════

async def login_droom(page, shot_dir: Path) -> bool:
    """D-Room にログインする（強制ログイン対応）"""
    print("── D-Room にログイン中 ──")
    await page.goto(DROOM_LOGIN_URL, wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(2)
    try:
        await page.locator('#txtEnterpriseId').fill(DROOM_TENPO)
        await page.locator('#txtUserId').fill(DROOM_TANTO)
        await page.locator('#txtPassword').fill(DROOM_PASS)
        await page.locator('#btnLoginButton').click()
    except Exception as e:
        print(f"  ⚠ D-Room ログインフォームエラー: {e}")
        return False

    try:
        await page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass
    await asyncio.sleep(2)

    # 強制ログイン（別セッションが存在する場合）
    if "CheckLogin" in page.url:
        print("  強制ログイン処理中...")
        try:
            await page.locator('#forcepassword').fill(DROOM_PASS)
        except Exception:
            pass
        try:
            await page.locator('#forceloginok').click()
        except Exception as e:
            print(f"  ⚠ 強制ログインボタンエラー: {e}")
            return False
        try:
            await page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass
        await asyncio.sleep(3)

    # お知らせページ → 「メニューに進む」クリック
    try:
        btn = page.get_by_text("メニューに進む").first
        if await btn.is_visible(timeout=3000):
            await btn.click()
            await asyncio.sleep(2)
    except Exception:
        pass

    await page.screenshot(path=str(shot_dir / "dr_01_after_login.png"))
    print(f"  D-Room ログイン後URL: {page.url}")

    if "login" in page.url.lower():
        print("  ⚠ D-Room ログイン失敗")
        return False
    print("  ✓ D-Room ログイン成功")
    return True


async def search_droom(page, area: str, rent_max: int, shot_dir: Path) -> list:
    """D-Room で物件を検索し、room_id 付き物件リストを返す"""
    print(f"── D-Room 物件検索: area={repr(area)} / {rent_max}円 ──")

    # エリアが空の場合は全件検索になり「最大検索件数」エラーになるためスキップ
    if not area.strip():
        print("  D-Room: エリア未指定のため検索をスキップします")
        return []

    await page.goto(DROOM_ROOMLIST_URL, wait_until="domcontentloaded", timeout=20000)
    await asyncio.sleep(3)

    # 住所入力
    try:
        await page.locator('input[name="address"]').fill(area)
        print(f"  住所入力: {area}")
    except Exception as e:
        print(f"  ⚠ 住所入力エラー: {e}")

    # 家賃上限（万円単位）
    rent_man = max(1, rent_max // 10000)
    try:
        await page.locator('input[name="rentTo"]').fill(str(rent_man))
        print(f"  家賃上限: {rent_man}万円")
    except Exception as e:
        print(f"  ⚠ 家賃上限入力エラー: {e}")

    # フォーム送信（address フィールドを持つフォームを優先して submit）
    await page.evaluate("""
        () => {
            const forms = Array.from(document.querySelectorAll('form'));
            const target = forms.find(f => f.querySelector('input[name="address"]')) || forms[0];
            if (target) target.submit();
        }
    """)
    await asyncio.sleep(5)
    try:
        await page.wait_for_load_state("domcontentloaded", timeout=15000)
    except Exception:
        pass
    await asyncio.sleep(2)

    await page.screenshot(path=str(shot_dir / "dr_02_search_result.png"))
    print(f"  D-Room 検索結果URL: {page.url}")

    # 最大検索件数エラーを検出 → エリアで絞り込めなかった場合は空リストを返す
    page_text = await page.evaluate("() => document.body.innerText")
    if '最大検索件数' in page_text:
        print(f"  ⚠ D-Room: 検索件数が上限を超えました（エリア '{area}' では絞り込めませんでした）")
        return []

    # 物件チェックボックス（value が「数字-数字-数字」形式）から物件情報を抽出
    # bukken_name 属性に建物名、前の行に住所・賃料が含まれている
    props = await page.evaluate("""
        () => {
            const results = [];
            const seen = new Set();
            for (const cb of document.querySelectorAll('input[type="checkbox"]')) {
                if (!cb.value || !cb.value.match(/^\\d{7,12}-\\d{3}/)) continue;
                if (seen.has(cb.value)) continue;
                seen.add(cb.value);

                // bukken_name 属性から建物名を取得
                const bukkenName = (cb.getAttribute('bukken_name') || cb.value).trim();
                const row = cb.closest('tr');
                let address = '', rent = '', layout = '', area = '';

                if (row) {
                    // 前の5行を調べて住所・賃料を探す
                    let prev = row.previousElementSibling;
                    for (let i = 0; i < 6 && prev; i++) {
                        const text = prev.innerText.replace(/\\s+/g, ' ').trim();
                        // 住所行（都道府県+丁目/番）
                        if (!address && text.match(/[都道府県市区町村]/) &&
                            (text.match(/\\d+丁目|\\d+番|\\d+-\\d+/) || text.match(/\\d+号/))) {
                            const addrPart = text.split(' - ')[0].split('　')[0].trim();
                            if (addrPart.length > 5) address = addrPart.substring(0, 60);
                        }
                        // 賃料行
                        if (!rent) {
                            const rentMatch = text.match(/(\\d{2,3},\\d{3})円/);
                            if (rentMatch) rent = rentMatch[0];
                        }
                        // 面積
                        if (!area) {
                            const sqmMatch = text.match(/(\\d+\\.?\\d*㎡)/);
                            if (sqmMatch) area = sqmMatch[0];
                        }
                        // 間取り
                        if (!layout) {
                            const layoutMatch = text.match(/([123][KLDK]+|ワンルーム|1R)/);
                            if (layoutMatch) layout = layoutMatch[0];
                        }
                        prev = prev.previousElementSibling;
                    }
                }
                results.push({
                    room_id: cb.value,
                    name: bukkenName.substring(0, 40),
                    address, rent, layout, area, station: ''
                });
            }
            return results;
        }
    """)

    print(f"  D-Room 物件数: {len(props)}")
    for p in props[:5]:
        print(f"    {p['name'][:25]} / {p.get('rent','')} / {p['room_id']}")

    for p in props:
        p['source'] = 'droom'

    return props


async def download_droom_bulk_pdf(page, out_dir: Path, max_props: int = 30) -> str:
    """D-Room の物件選択チェックボックスをONにしてDroomシート一括PDFをダウンロードする"""
    print("── D-Room Droomシート一括出力 ──")

    # 物件行チェックボックスのみを選択（フィルタ用CBは除外）
    selected = await page.evaluate("""
        (maxProps) => {
            let cnt = 0;
            for (const cb of document.querySelectorAll('input[type="checkbox"]')) {
                if (!cb.value || !cb.value.match(/^\\d{7,12}-\\d{3}/)) continue;
                if (cb.disabled) continue;
                cb.checked = true;
                cb.dispatchEvent(new Event('change', { bubbles: true }));
                cb.dispatchEvent(new Event('click',  { bubbles: true }));
                cnt++;
                if (cnt >= maxProps) break;
            }
            return cnt;
        }
    """, max_props)
    print(f"  選択: {selected}件")
    await asyncio.sleep(1)

    if selected == 0:
        print("  ⚠ 選択できる物件がありません")
        return ""

    # 一括ダウンロードボタンをクリック → PDF ダウンロードを待機
    try:
        async with page.expect_download(timeout=90000) as dl_info:
            await page.locator('#btn-yikkatudownload').click()
        dl = await dl_info.value
        fname = dl.suggested_filename or "D-Room一括出力.pdf"
        save_path = str(out_dir / fname)
        await dl.save_as(save_path)
        size = os.path.getsize(save_path)
        print(f"  ✓ D-Room一括PDF: {fname} ({size:,} bytes)")
        return save_path
    except Exception as e:
        print(f"  ⚠ D-Room一括PDFダウンロードエラー: {e}")
        return ""


# ══════════════════════════════════════════════════════════
#  レオパレス21 検索・PDF出力（ログイン不要・一般サイト）
# ══════════════════════════════════════════════════════════

LP_BASE = "https://www.leopalace21.com"

async def search_leopalace(page, area: str, rent_max: int, shot_dir: Path,
                           work_address: str = '') -> list:
    """レオパレス21 一般サイトで物件を検索してリストを返す"""
    print(f"── レオパレス21 物件検索: {area} / {rent_max}円 ──")

    # 都道府県コードを推定（city名だけでは不明な場合は work_address で補う）
    pref_code = None
    for src in [area, work_address]:
        if not src:
            continue
        for pref_name, code in PREF_CODE.items():
            if pref_name in src:
                pref_code = code
                break
        if pref_code:
            break
    if not pref_code:
        print(f"  ⚠ 都道府県コード不明: {area}")
        return []

    # 都道府県ページから対象エリアのURLを取得
    pref_url = f"{LP_BASE}/search/chintai/area?prefectureCode={pref_code}"
    await page.goto(pref_url, wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(4)
    await page.screenshot(path=str(shot_dir / "lp_01_pref.png"))

    # エリア名でA linkを探す（例: "大阪市北区"）
    area_url = await page.evaluate(f"""
        () => {{
            const links = Array.from(document.querySelectorAll('a[href]'));
            const m = links.find(a =>
                (a.innerText || '').trim() === '{area}' ||
                (a.innerText || '').trim().startsWith('{area}')
            );
            return m ? m.href : null;
        }}
    """)
    print(f"  エリアURL: {area_url}")

    if not area_url:
        print(f"  ⚠ エリアURL見つからず: {area}")
        return []

    # エリア物件一覧へ
    await page.goto(area_url, wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(4)
    await page.screenshot(path=str(shot_dir / "lp_02_area_list.png"))

    # 家賃上限フィルタをセット（rentTo）
    # 値リスト: 20000,25000,...,80000,...
    try:
        # 近い値を選ぶ: rent_max 以下で最大の選択肢
        rent_options = await page.evaluate("""
            () => Array.from((document.querySelector('select#rentTo')||{options:[]}).options)
                .map(o => ({ value: o.value, text: o.text }))
                .filter(o => o.value && !isNaN(o.value))
        """)
        best_val = ""
        best_int = 0
        for opt in rent_options:
            v = int(opt['value'])
            if v <= rent_max and v > best_int:
                best_int = v
                best_val = opt['value']
        if best_val:
            await page.select_option('select#rentTo', best_val)
            print(f"  家賃上限: {best_val}円に設定")
            await asyncio.sleep(1)
    except Exception as e:
        print(f"  家賃フィルタスキップ: {e}")

    # 「検索（N件）」ボタンをクリック
    try:
        search_btn = page.locator("button").filter(has_text=re.compile(r"検索（\d+件）")).first
        if await search_btn.is_visible(timeout=5000):
            await search_btn.click()
            await asyncio.sleep(4)
        await page.screenshot(path=str(shot_dir / "lp_03_results.png"))
        print(f"  検索後URL: {page.url}")
    except Exception as e:
        print(f"  検索ボタンエラー: {e}")

    # 物件カード（建物単位）のリンクを収集
    prop_links = await page.evaluate("""
        () => {
            const seen = new Set();
            return Array.from(document.querySelectorAll('a[href]'))
                .filter(a => a.href.includes('/properties/common/'))
                .map(a => {
                    const card = a.closest('article, li, section') || a.parentElement;
                    const rawText = (card ? card.innerText : a.innerText || '').trim();
                    // 行単位で解析
                    const lines = rawText.split('\\n').map(l => l.trim()).filter(l => l);
                    // 建物名: 「オンライン相談・契約」以外で最初の行
                    const nameIdx = lines.findIndex(l =>
                        !l.includes('相談') && !l.includes('契約') && !l.includes('オンライン'));
                    const name = nameIdx >= 0 ? lines[nameIdx] : '';
                    // 住所: 都道府県を含む行
                    const address = lines.find(l => l.match(/[都道府県]/)) || '';
                    // 家賃: 数字のみの行+次行が「万円」
                    let rent = '';
                    for (let i = 0; i < lines.length - 1; i++) {
                        if (lines[i].match(/^\\d+\\.?\\d*$/) && lines[i+1].trim() === '万円') {
                            rent = lines[i] + '万円';
                            break;
                        }
                        if (lines[i].match(/^\\d+\\.?\\d*万円/)) {
                            rent = lines[i].match(/^(\\d+\\.?\\d*万円)/)[1];
                            break;
                        }
                    }
                    // 間取り: 1K/1R等
                    const layoutLine = lines.find(l => l.match(/^[1-4][RKLDK]+$/));
                    // 面積
                    const areaLine = lines.find(l => l.match(/^\\d+\\.\\d+㎡$/));
                    return {
                        href: a.href,
                        name: name.substring(0, 40),
                        address: address.substring(0, 60),
                        rent: rent,
                        layout: layoutLine || '',
                        area: areaLine || ''
                    };
                })
                .filter(x => {
                    if (seen.has(x.href)) return false;
                    seen.add(x.href);
                    return true;
                });
        }
    """)
    print(f"  建物カード: {len(prop_links)}件")

    results = []
    for pl in prop_links[:5]:
        results.append({
            'name': pl['name'] or pl['href'].split('/')[-1][:30],
            'rent': pl['rent'],
            'layout': pl['layout'],
            'area': pl['area'],
            'address': pl['address'] or area,
            'detail_url': pl['href'],
            'station': '',
        })
        print(f"    {results[-1]['name'][:25]} / {results[-1]['rent']} / {results[-1]['address'][:30]}")

    print(f"  レオパレス: {len(results)}件取得")
    return results


async def download_leopalace_pdf(page, playwright_instance,
                                  detail_url: str, prop_name: str,
                                  out_dir: Path) -> str:
    """レオパレス物件詳細ページをPDF化して保存"""
    prop_safe = re.sub(r'[\\/:*?"<>|]', '_', prop_name)[:50]
    out_pdf = str(out_dir / f"{prop_safe}_物件詳細.pdf")
    try:
        # ヘッドレスブラウザで高品質PDF
        cookies = await page.context.cookies()
        hb = await playwright_instance.chromium.launch(headless=True)
        hc = await hb.new_context(viewport={"width": 1280, "height": 900}, locale="ja-JP")
        await hc.add_cookies(cookies)
        hp = await hc.new_page()
        await hp.goto(detail_url, wait_until="domcontentloaded", timeout=30000)
        await asyncio.sleep(3)
        await hp.pdf(path=out_pdf, format="A4", print_background=True,
                     margin={"top": "10mm", "bottom": "10mm", "left": "10mm", "right": "10mm"})
        await hb.close()
        size = os.path.getsize(out_pdf)
        print(f"  ✓ レオパレスPDF: {out_pdf} ({size:,} bytes)")
        return out_pdf
    except Exception as e:
        print(f"  ⚠ レオパレスPDF失敗: {e}")
        # フォールバック: 通常ページでPDF
        try:
            await page.goto(detail_url, wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(3)
            await page.pdf(path=out_pdf, format="A4", print_background=True)
            print(f"  ✓ レオパレスPDF(fallback): {out_pdf}")
            return out_pdf
        except Exception as e2:
            print(f"  ⚠ レオパレスPDFフォールバック失敗: {e2}")
            return ""


# ══════════════════════════════════════════════════════════
#  リアブロ (リアネットプロ) ログイン・検索・PDF出力
# ══════════════════════════════════════════════════════════

async def login_reabro(page, shot_dir: Path) -> bool:
    """リアブロ（リアネットプロ）にログインする"""
    print("── リアブロ にログイン中 ──")
    await page.goto(REABRO_LOGIN_URL, wait_until="domcontentloaded", timeout=30000)
    await asyncio.sleep(2)
    try:
        await page.locator('input[name="id"]').fill(REABRO_ID)
        await page.locator('input[name="pass"]').fill(REABRO_PASS)
        await page.locator('button:has-text("ログイン")').first.click()
    except Exception as e:
        print(f"  ⚠ リアブロ ログインフォームエラー: {e}")
        return False

    try:
        await page.wait_for_load_state("domcontentloaded", timeout=15000)
    except Exception:
        pass
    await asyncio.sleep(2)
    await page.screenshot(path=str(shot_dir / "rb_01_after_login.png"))
    print(f"  リアブロ ログイン後URL: {page.url}")

    if "main.php" in page.url or "estate" in page.url or "realnetpro" in page.url:
        if "index.php" not in page.url:
            print("  ✓ リアブロ ログイン成功")
            return True
    print("  ⚠ リアブロ ログイン失敗")
    return False


async def search_reabro(page, area: str, rent_max: int, shot_dir: Path,
                        work_address: str = '') -> list:
    """リアブロ で物件を検索し、room_id 付き物件リストを返す。

    検索フロー:
      1. 建物一覧ページへ遷移
      2. 就業先住所から都道府県コードを判定 → pref_code チェックボックスをクリック
      3. AJAX で都市リスト更新後、city_code[] から市区町村名に一致するラベルをクリック
      4. rental_cost2 で家賃上限を設定
      5. 「検索」ボタンをクリック → 結果を取得
    """
    print(f"── リアブロ 物件検索: area={repr(area)} / {rent_max}円 ──")

    # エリア・勤務地住所の両方が空の場合は都市フィルターなし検索になるためスキップ
    if not area.strip() and not work_address.strip():
        print("  リアブロ: エリア未指定のため検索をスキップします")
        return []

    # ── JS: .room_system_menu[title] から物件情報を抽出 ──────
    EXTRACT_ROOMS_JS = """
        () => {
            const results = [];
            document.querySelectorAll('.room_system_menu[title]').forEach(el => {
                const t = el.getAttribute('title');
                if (!t || !t.includes(',')) return;
                const parts = t.split(',');
                const room_id = parts[0].trim();
                if (!room_id || !/^\\d+$/.test(room_id)) return;
                const row = el.closest('tr') || el.closest('li') || el.parentElement;
                const text = row ? row.innerText.replace(/\\s+/g,' ').trim() : '';
                let rent = '', layout = '', sq = '';
                const rentM = text.match(/([\\d,]+\\.?\\d*)(万円|円)/);
                if (rentM) rent = rentM[0];
                const layoutM = text.match(/([1-9][LKDS]+|ワンルーム)/i);
                if (layoutM) layout = layoutM[0];
                const sqM = text.match(/\\d+\\.?\\d*㎡/);
                if (sqM) sq = sqM[0];
                results.push({
                    room_id:       room_id,
                    building_name: (parts[1] || '').trim(),
                    room_name:     (parts[2] || '').trim(),
                    rent: rent, layout: layout, area: sq
                });
            });
            return results;
        }
    """

    # 建物一覧ページへ
    estate_url = REABRO_BASE_URL + "/main.php?method=estate&display=building"
    await page.goto(estate_url, wait_until="networkidle", timeout=30000)
    await asyncio.sleep(3)
    await page.screenshot(path=str(shot_dir / "rb_02_estate_list.png"))

    # ── Step1: 所在地絞り込みモーダルで都道府県 → 市区町村を選択 ────
    #
    # 確認済みフロー (debug_reabro_step2.py で検証):
    #   A. アコーディオンクリック → step2 (市区郡選択, デフォルト東京) が開く
    #   B. step2 内の「都道府県の設定」(.step1_m_text) をクリック → step1 (都道府県) へ
    #   C. label.one_pref[value=XX] をクリック → 都道府県を選択
    #   D. .next_step_button.next_action をクリック → step2 (選択済み都道府県の市区郡)
    #   E. step2 内で市区町村ラベルをクリック
    #   F. ×とじる でモーダルを閉じる
    # ─────────────────────────────────────────────────────────────
    target_pref_code = None
    if work_address:
        for pref_name, code in PREF_CODE.items():
            if pref_name in work_address:
                target_pref_code = code
                break

    city_selected = False
    if target_pref_code:
        try:
            # A: アコーディオンを開く → step2 (東京のデフォルト) 表示
            slide_btn = page.locator('.one_slide_search_box .click_menu').first
            if await slide_btn.count() > 0:
                await slide_btn.click()
                await asyncio.sleep(2)
                print("  所在地パネル(step2) opened")

            # B: 「都道府県の設定」→ step1 を開く
            await page.evaluate(
                "() => { const el = document.querySelector('.step1_m_text'); if (el) el.click(); }"
            )
            await asyncio.sleep(1.5)

            # C: 目標の都道府県ラベルをクリック（step1 で visible になる）
            pref_label = page.locator(f'label.one_pref:has(input[value="{target_pref_code}"])')
            pref_cnt = await pref_label.count()
            if pref_cnt > 0:
                await pref_label.first.click()
                await asyncio.sleep(1.5)
                print(f"  都道府県クリック: pref_code={target_pref_code}")
            else:
                # フォールバック: evaluate でラベルクリック
                await page.evaluate(f"""
                    () => {{
                        const lbl = document.querySelector(
                            'label.one_pref:has(input[value="{target_pref_code}"])');
                        if (lbl) lbl.click();
                    }}
                """)
                await asyncio.sleep(1.5)
                print(f"  都道府県クリック(evaluate): pref_code={target_pref_code}")

            # D: 「市区郡の設定へ進む」(.next_step_button.next_action) → step2 更新
            next_btn = page.locator('.next_step_button.next_action')
            if await next_btn.count() > 0:
                await next_btn.first.click()
                await asyncio.sleep(2)
                print("  市区郡の設定へ進む クリック")
            await page.screenshot(path=str(shot_dir / "rb_03_pref_clicked.png"))

            # E: step2 内で市区町村ラベルをクリック
            for city_query in [area, re.sub(r'[市区町村郡]$', '', area)]:
                found = await page.evaluate(f"""
                    () => {{
                        const target = "{city_query}";
                        const s2 = document.querySelector('.step2.city_select.detail_select_box');
                        const labels = Array.from(
                            s2 ? s2.querySelectorAll('label') : document.querySelectorAll('label')
                        );
                        const lbl = labels.find(l => l.textContent.trim().includes(target));
                        if (lbl) {{
                            lbl.click();
                            return {{found: true, text: lbl.textContent.trim().substring(0, 20)}};
                        }}
                        return {{found: false}};
                    }}
                """)
                if found.get('found'):
                    city_selected = True
                    await asyncio.sleep(1)
                    print(f"  都市選択: {found['text']}")
                    break

            if not city_selected:
                print(f"  ⚠ 都市 '{area}' のラベルが見つかりません（都道府県全体で検索します）")

            # F: ×とじる でモーダルを閉じる
            await page.evaluate("""
                () => {
                    const spans = Array.from(document.querySelectorAll('span'));
                    const close = spans.find(s =>
                        s.textContent.trim() === '×とじる' &&
                        s.getBoundingClientRect().width > 0
                    );
                    if (close) close.click();
                }
            """)
            await asyncio.sleep(1)
            await page.screenshot(path=str(shot_dir / "rb_03b_city_selected.png"))

        except Exception as e:
            print(f"  ⚠ 所在地選択エラー: {e}")

    # ── Step3: 家賃上限を設定 ────────────────────────────────
    RENT_STEPS = [
        20000, 25000, 30000, 35000, 40000, 45000, 50000, 55000,
        60000, 65000, 70000, 75000, 80000, 85000, 90000, 95000,
        100000, 110000, 120000, 130000, 140000, 150000,
        160000, 170000, 180000, 190000, 200000,
    ]
    best_rent_val = '-1'  # 上限なし
    for v in RENT_STEPS:
        if v <= rent_max:
            best_rent_val = str(v)
    try:
        await page.select_option('select[name="rental_cost2"]', best_rent_val)
        disp = f'{int(best_rent_val)//10000}万円' if best_rent_val != '-1' else '上限なし'
        print(f"  家賃上限設定: {disp}（rent_max={rent_max}円）")
    except Exception as e:
        print(f"  ⚠ 家賃上限設定エラー: {e}")

    # ── Step4: 検索ボタンをクリック ───────────────────────────
    try:
        # まず visible な 検索ボタンを evaluate で探してクリック
        clicked_search = await page.evaluate("""
            () => {
                const candidates = Array.from(
                    document.querySelectorAll('button, input[type="button"], input[type="submit"]')
                );
                const btn = candidates.find(el => {
                    const t = (el.innerText || el.value || '').trim();
                    const r = el.getBoundingClientRect();
                    return t === '検索' && r.width > 0 && r.height > 0;
                });
                if (btn) { btn.click(); return true; }
                return false;
            }
        """)
        if not clicked_search:
            # フォールバック: フォームを直接 submit
            await page.evaluate(
                "() => { document.getElementById('main_form')?.submit(); }"
            )
            print("  検索: フォーム submit (fallback)")
        try:
            await page.wait_for_load_state("networkidle", timeout=12000)
        except Exception:
            pass
        await asyncio.sleep(4)
        await page.screenshot(path=str(shot_dir / "rb_04_search_result.png"))
        print(f"  検索実行後URL: {page.url}")
    except Exception as e:
        print(f"  ⚠ 検索ボタンエラー: {e}")

    # ── Step5: 結果取得 ───────────────────────────────────────
    area_rooms = await page.evaluate(EXTRACT_ROOMS_JS)
    print(f"  エリア絞り込み後: {len(area_rooms)}件")

    # 0件の場合はキーワード検索でフォールバック
    if not area_rooms and area:
        print(f"  → キーワード検索フォールバック: {area}")
        try:
            kw = page.locator('input[name="keyword"]').first
            if await kw.is_visible(timeout=2000):
                await kw.click(click_count=3)
                await kw.fill(area)
                await kw.press("Enter")
                try:
                    await page.wait_for_load_state("networkidle", timeout=10000)
                except Exception:
                    pass
                await asyncio.sleep(3)
                area_rooms = await page.evaluate(EXTRACT_ROOMS_JS)
                print(f"  キーワード検索({area}): {len(area_rooms)}件")
        except Exception as e:
            print(f"  キーワード検索エラー: {e}")

    # 物件リストに変換
    result = []
    for r in area_rooms:
        name = r['building_name'] or 'リアブロ物件'
        if r['room_name']:
            name = f"{name} {r['room_name']}"
        result.append({
            'room_id': r['room_id'],
            'name':    name[:40],
            'address': r.get('address', ''),
            'rent':    r.get('rent', ''),
            'layout':  r.get('layout', ''),
            'area':    r.get('area', ''),
            'station': '',
            'source':  'reabro',
        })

    print(f"  リアブロ 物件数: {len(result)}")
    for p in result[:5]:
        print(f"    {p['name'][:30]} / {p.get('rent', '')} / room_id={p['room_id']}")

    return result


async def _download_reabro_pdf_url(ctx_or_page, url: str, save_path: str) -> bool:
    """リアブロのPDF URLをCookieを引き継いでurllibでダウンロードする"""
    import urllib.request as _urllib_req
    try:
        # ctx (BrowserContext) または page から cookies を取得
        if hasattr(ctx_or_page, 'cookies'):
            cookies = await ctx_or_page.cookies(urls=[REABRO_BASE_URL])
        else:
            cookies = await ctx_or_page.context.cookies()
        cookie_str = "; ".join(f"{c['name']}={c['value']}" for c in cookies)
        req = _urllib_req.Request(url, headers={
            "Cookie":     cookie_str,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Referer":    REABRO_BASE_URL + "/main.php",
        })
        with _urllib_req.urlopen(req, timeout=30) as r:
            content = r.read()
        if not content.startswith(b"%PDF"):
            print(f"    ⚠ PDFヘッダーなし: {url[:60]}")
            return False
        with open(save_path, "wb") as f:
            f.write(content)
        print(f"    ✓ ダウンロード: {os.path.basename(save_path)} ({len(content):,} bytes)")
        return True
    except Exception as e:
        print(f"    ⚠ URLダウンロードエラー ({url[:60]}): {e}")
        return False


async def download_reabro_pdfs(ctx, room_id: str, prop_name: str,
                               out_dir: Path, font_name: str,
                               map_img_path: str = "",
                               commute_method: str = '',
                               workplace: str = '') -> dict:
    """リアブロ物件の社宅.com版(客付け)・元図面(元付け)PDFをダウンロードし地図付きを生成する。
    手順書より:
      客付け資料 (社宅.com版/オレンジ) = /common/factsheet.php?id={id}
      元付け資料 (元図面/緑)           = /common/factsheet.php?id={id}&org=2
    """
    print(f"  リアブロ PDF取得: id={room_id} / {prop_name[:20]}")
    prop_safe = re.sub(r'[\\/:*?"<>|]', "_", prop_name)[:40]
    result = {"syataku": "", "mototsuke": "", "merged": ""}

    syataku_url   = f"{REABRO_BASE_URL}/common/factsheet.php?id={room_id}"          # 社宅.com版
    mototsuke_url = f"{REABRO_BASE_URL}/common/factsheet.php?id={room_id}&org=2"    # 元図面

    syataku_path   = str(out_dir / f"{prop_safe}_社宅com版.pdf")
    mototsuke_path = str(out_dir / f"{prop_safe}_元図面.pdf")

    # ① 社宅.com版（客付け資料）
    ok = await _download_reabro_pdf_url(ctx, syataku_url, syataku_path)
    if ok:
        result["syataku"] = syataku_path

    # ② 元図面（元付け資料）
    ok = await _download_reabro_pdf_url(ctx, mototsuke_url, mototsuke_path)
    if ok:
        result["mototsuke"] = mototsuke_path

    # ③ 社宅.com版 + 元図面 + 地図ページ をマージして合本PDF を生成
    pdfs_to_merge = [p for p in [result["syataku"], result["mototsuke"]] if p and os.path.exists(p)]
    if pdfs_to_merge:
        merged_path = str(out_dir / f"{prop_safe}_地図付き.pdf")
        try:
            import io as _io
            writer = PdfWriter()
            for pdf_path in pdfs_to_merge:
                for pg in PdfReader(pdf_path).pages:
                    writer.add_page(pg)

            # 地図ページ追加
            if map_img_path and os.path.exists(map_img_path):
                W, H = A4
                M = 15 * mm
                map_buf = _io.BytesIO()
                cv = canvas.Canvas(map_buf, pagesize=A4)
                cv.setFont(font_name, 11)
                cv.setFillColor(colors.HexColor("#37474f"))
                header_parts = ["通勤ルート地図"]
                if commute_method:
                    header_parts.append(f"（{commute_method}）")
                if workplace:
                    short_wp = workplace[:30] + ('…' if len(workplace) > 30 else '')
                    header_parts.append(f"  勤務地: {short_wp}")
                cv.drawString(M, H - M, ''.join(header_parts))
                img = Image.open(map_img_path)
                iw, ih = img.size
                scale = min((W - 2 * M) / iw, (H - 2 * M - 15 * mm) / ih)
                dw, dh = iw * scale, ih * scale
                cv.drawImage(map_img_path, M, H - M - 15 * mm - dh,
                             width=dw, height=dh)
                cv.save()
                map_buf.seek(0)
                for pg in PdfReader(map_buf).pages:
                    writer.add_page(pg)

            with open(merged_path, "wb") as f:
                writer.write(f)
            result["merged"] = merged_path
            print(f"    ✓ 地図付きPDF: {merged_path}")
        except Exception as e:
            print(f"    ⚠ リアブロ PDF合成エラー: {e}")
            import traceback; traceback.print_exc()

    return result


def parse_property_text(text: str, idx: int) -> dict:
    """物件テキストから情報をパースする（簡易）"""
    import re
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    prop = {
        "name": lines[0] if lines else f"物件{idx}",
        "address": "",
        "rent": "",
        "layout": "",
        "area": "",
        "age": "",
        "station": "",
        "raw": text[:300],
    }

    for line in lines:
        if re.search(r"[都道府県市区町村]", line) and not prop["address"]:
            prop["address"] = line
        if re.search(r"\d+,?\d*円|万円", line) and not prop["rent"]:
            prop["rent"] = line
        if re.search(r"\dK|\dLDK|\dDK|ワンルーム", line) and not prop["layout"]:
            prop["layout"] = re.search(r"(\d[KLDK]+|ワンルーム)", line).group()
        if re.search(r"\d+\.?\d*㎡|m²|平米", line) and not prop["area"]:
            prop["area"] = line
        if re.search(r"築\d+年|新築", line) and not prop["age"]:
            prop["age"] = re.search(r"(築\d+年|新築)", line).group()
        if re.search(r"駅|バス停", line) and not prop["station"]:
            prop["station"] = line

    return prop


# ══════════════════════════════════════════════════════════
#  Google マップ 通勤ルート
# ══════════════════════════════════════════════════════════

def _commute_travel_mode(commute_method: str) -> str:
    """
    通勤方法文字列を Google Maps travelmode パラメータに変換する。
    自転車→bicycling / 電車・バス→transit / 車・自動車→driving / 徒歩→walking
    """
    m = commute_method or ''
    if re.search(r'自転車|チャリ', m):
        return 'bicycling'
    if re.search(r'車|自動車|カー', m):
        return 'driving'
    if re.search(r'電車|バス|鉄道|公共', m):
        return 'transit'
    if re.search(r'徒歩|歩き', m):
        return 'walking'
    # デフォルト: 電車
    return 'transit'


async def get_maps_screenshot(page, from_addr: str, to_addr: str, out_path: str,
                               commute_method: str = '') -> bool:
    """
    Google マップで物件住所→勤務地住所の通勤ルートをスクリーンショット保存する。

    Parameters
    ----------
    from_addr       : 物件住所（出発地）
    to_addr         : 勤務地住所（目的地）
    out_path        : 保存先パス（.png）
    commute_method  : 通勤方法（自転車/電車/車 など）→ travelmode に自動変換
    """
    from urllib.parse import quote
    travel_mode = _commute_travel_mode(commute_method)
    print(f"  地図取得: {from_addr[:30]} → {to_addr[:30]}  [{travel_mode}]")
    url = (
        f"https://www.google.com/maps/dir/?api=1"
        f"&origin={quote(from_addr)}"
        f"&destination={quote(to_addr)}"
        f"&travelmode={travel_mode}"
    )
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await asyncio.sleep(3)

        # Google の同意画面・Cookie バナーを閉じる
        for btn_text in ["同意する", "すべて同意", "I agree", "Accept all", "Agree"]:
            try:
                await page.click(f'button:has-text("{btn_text}")', timeout=1500)
                await asyncio.sleep(1)
                break
            except Exception:
                pass

        # ルートが描画されるまで待機（経路パネルが表示されたら OK）
        try:
            await page.wait_for_selector(
                '[data-panel-id="directions"], .section-directions-trip, .directions-travel-mode-selector',
                timeout=10000
            )
        except Exception:
            pass
        await asyncio.sleep(3)

        await page.screenshot(path=out_path, full_page=False)
        print(f"    ✓ 地図保存: {out_path}")
        return True
    except Exception as e:
        print(f"    ⚠ 地図取得エラー: {e}")
        return False


# ══════════════════════════════════════════════════════════
#  PDF 生成
# ══════════════════════════════════════════════════════════

def generate_pdf(case: dict, properties: list[dict], map_shots: dict,
                 font_name: str, out_path: Path):
    print(f"── PDF生成: {out_path} ──")
    c = canvas.Canvas(str(out_path), pagesize=A4)
    W, H = A4
    M = 20 * mm

    def new_page():
        c.showPage()
        return H - M

    def draw_text(text, x, y, size=10, color=colors.black, bold=False):
        c.setFont(font_name, size)
        c.setFillColor(color)
        c.drawString(x, y, text)

    # ── タイトルページ ──────────────────────────────────
    y = H - M

    c.setFont(font_name, 20)
    c.setFillColor(colors.HexColor("#1a237e"))
    c.drawString(M, y, "社宅物件提案書")
    y -= 8 * mm

    c.setStrokeColor(colors.HexColor("#1a237e"))
    c.setLineWidth(1.5)
    c.line(M, y, W - M, y)
    y -= 8 * mm

    # 依頼者情報
    c.setFont(font_name, 12)
    c.setFillColor(colors.HexColor("#37474f"))
    c.drawString(M, y, "■ 依頼内容")
    y -= 7 * mm

    info_items = [
        ("氏名",        case.get("氏名", "")),
        ("入居希望日",   str(case.get("入居希望日", ""))),
        ("転勤元",      case.get("転勤元住所", "")),
        ("勤務先住所",   case.get("勤務先住所（転勤先）", case.get("勤務先住所", ""))),
        ("希望エリア",   case.get("希望エリア", "")),
        ("希望社宅種類", case.get("希望社宅種類", "")),
        ("家賃上限",     f"{case.get('家賃上限（円）', case.get('家賃上限', ''))}円"),
        ("希望間取り",   case.get("希望間取り", "")),
        ("通勤方法",     f"{case.get('通勤方法', '')} / {case.get('通勤可能時間', '')}"),
    ]

    c.setFont(font_name, 10)
    for label, val in info_items:
        c.setFillColor(colors.HexColor("#546e7a"))
        c.drawString(M, y, f"  {label}:")
        c.setFillColor(colors.black)
        c.drawString(M + 32 * mm, y, str(val))
        y -= 6 * mm

    y -= 5 * mm

    # ── 物件候補 ─────────────────────────────────────────
    c.setFont(font_name, 12)
    c.setFillColor(colors.HexColor("#37474f"))
    num_props = len(properties[:4])
    c.drawString(M, y, f"■ 物件候補（{num_props}件）")
    y -= 8 * mm

    workplace = case.get("勤務先住所（転勤先）", case.get("勤務先住所", ""))

    for i, prop in enumerate(properties[:4], 1):
        # ページ余白チェック
        if y < 60 * mm:
            y = new_page()

        # 物件ヘッダー
        c.setFillColor(colors.HexColor("#e3f2fd"))
        c.rect(M - 2*mm, y - 2*mm,
               W - 2*M + 4*mm, 10*mm, fill=1, stroke=0)
        c.setFont(font_name, 11)
        c.setFillColor(colors.HexColor("#0d47a1"))
        c.drawString(M, y + 2*mm, f"物件 {i}")
        y -= 10 * mm

        # 物件詳細
        details = [
            ("物件名",   prop.get("name", "")),
            ("住所",     prop.get("address", "")),
            ("家賃",     prop.get("rent", "")),
            ("間取り",   prop.get("layout", "")),
            ("面積",     prop.get("area", "")),
            ("築年数",   prop.get("age", "")),
            ("最寄駅",   prop.get("station", "")),
        ]
        c.setFont(font_name, 10)
        for label, val in details:
            if val:
                c.setFillColor(colors.HexColor("#546e7a"))
                c.drawString(M + 3*mm, y, f"{label}:")
                c.setFillColor(colors.black)
                c.drawString(M + 22*mm, y, str(val)[:60])
                y -= 5.5 * mm

        y -= 3 * mm

        # 物件写真 / 間取り図
        madori_path = prop.get("madori_path", "")
        photo_label = prop.get("photo_label", "物件写真")
        if madori_path and os.path.exists(madori_path):
            if y < 80 * mm:
                y = new_page()
            c.setFont(font_name, 10)
            c.setFillColor(colors.HexColor("#37474f"))
            c.drawString(M, y, f"  {photo_label}")
            y -= 5 * mm
            try:
                img = Image.open(madori_path)
                iw, ih = img.size
                max_w = (W - 2 * M) * 0.5   # ページ幅の半分に収める
                max_h = 60 * mm
                scale = min(max_w / iw, max_h / ih)
                dw, dh = iw * scale, ih * scale
                if y - dh < 15 * mm:
                    y = new_page()
                c.drawImage(madori_path, M, y - dh, width=dw, height=dh)
                y -= dh + 8 * mm
            except Exception as e:
                print(f"  間取り図挿入エラー: {e}")
                y -= 3 * mm

        # 通勤ルート地図
        map_path = map_shots.get(i)
        if map_path and os.path.exists(map_path):
            if y < 120 * mm:
                y = new_page()

            c.setFont(font_name, 10)
            c.setFillColor(colors.HexColor("#37474f"))
            c.drawString(M, y, f"  通勤ルート地図  ({case.get('通勤方法', '')} / {case.get('通勤可能時間', '')})")
            y -= 5 * mm

            try:
                img = Image.open(map_path)
                iw, ih = img.size
                max_w = W - 2 * M
                max_h = 80 * mm
                scale = min(max_w / iw, max_h / ih)
                dw, dh = iw * scale, ih * scale
                if y - dh < 15 * mm:
                    y = new_page()
                c.drawImage(map_path, M, y - dh, width=dw, height=dh)
                y -= dh + 10 * mm
            except Exception as e:
                print(f"  画像挿入エラー: {e}")
                y -= 5 * mm
        else:
            y -= 5 * mm

        # 区切り線
        if i < len(properties[:4]):
            c.setStrokeColor(colors.HexColor("#cfd8dc"))
            c.setLineWidth(0.5)
            c.line(M, y, W - M, y)
            y -= 8 * mm

    # フッター
    c.setFont(font_name, 8)
    c.setFillColor(colors.grey)
    ts = datetime.now().strftime("%Y年%m月%d日 %H:%M 作成")
    c.drawString(M, 10 * mm, ts)

    c.save()
    print(f"  ✓ PDF保存: {out_path}")


# ══════════════════════════════════════════════════════════
#  メイン処理
# ══════════════════════════════════════════════════════════

async def process_case(playwright, case: dict, case_num: int, font_name: str):
    browser = await playwright.chromium.launch(
        headless=False,
        slow_mo=400,
        channel="chrome",
    )
    ctx = await browser.new_context(
        viewport={"width": 1280, "height": 900},
        locale="ja-JP",
    )
    page = await ctx.new_page()

    shot_dir = SHOT_DIR / f"case_{case_num:02d}"
    shot_dir.mkdir(exist_ok=True)

    try:
        area = case.get("希望エリア", "")
        rent_max = int(case.get("家賃上限（円）", case.get("家賃上限", 100000)))
        layout = case.get("希望間取り", "")
        properties = []

        # 管理番号を取得（なければ自動採番）
        case_id = case.get("管理番号", f"A{case_num:03d}")
        case_id = re.sub(r'[\\/:*?"<>|]', "_", str(case_id)).strip()
        # ケース別出力ディレクトリ
        case_dir = OUTPUT_DIR / case_id
        case_dir.mkdir(parents=True, exist_ok=True)
        print(f"  管理番号: {case_id}  →  {case_dir}")

        # ① ATBBログイン＆物件検索
        atbb_props = []
        atbb_detail_pdfs = {}   # atbb_idx → pdf_path
        try:
            logged_in = await login_atbb(page, shot_dir)
            if logged_in:
                atbb_props, page = await search_atbb(
                    page, ctx,
                    area=area,
                    rent_max=rent_max,
                    layout=layout,
                    shot_dir=shot_dir,
                )
                print(f"  ATBB: {len(atbb_props)}件取得")

                # 各物件に source/atbb_idx を付与
                for idx, ap in enumerate(atbb_props):
                    ap['source']   = 'atbb'
                    ap['atbb_idx'] = idx

                # 印刷(画像付き)PDFをダウンロード（case_dir に保存）
                for idx in range(min(len(atbb_props), 3)):
                    dp = await download_atbb_print_pdf(
                        page, ctx, playwright,
                        idx, atbb_props[idx], case_dir, font_name)
                    if dp:
                        atbb_detail_pdfs[idx] = dp
                        print(f"  ATBB PDF{idx+1}: {dp}")
        except Exception as e:
            print(f"  ATBB検索エラー: {e}")
            import traceback; traceback.print_exc()

        # ② 東建ルームサーチログイン＆物件検索（同一ブラウザの新規タブ）
        hm_props = []
        hm_detail_pdfs = {}   # hm_idx → pdf_path
        try:
            hm_page = await ctx.new_page()
            hm_shot_dir = shot_dir / "homemate"
            hm_shot_dir.mkdir(exist_ok=True)

            hm_logged_in = await login_homemate(hm_page, hm_shot_dir)
            if hm_logged_in:
                ok = await search_homemate(hm_page, area, hm_shot_dir,
                                           work_address=case.get('勤務地住所', ''))
                if ok:
                    hm_props = await extract_homemate_properties(hm_page, ctx, hm_shot_dir)
                    print(f"  東建: {len(hm_props)}件取得")

                    # 物件詳細PDFをダウンロード（hm_page を閉じる前に実行）
                    for idx, hp in enumerate(hm_props[:3]):
                        if hp.get('detail_href'):
                            dp = await download_homemate_detail_pdf(
                                hm_page, playwright, hp['detail_href'], case_dir)
                            hm_detail_pdfs[idx] = dp
                            print(f"  東建PDF{idx+1}: {dp}")
            await hm_page.close()
        except Exception as e:
            print(f"  東建検索エラー: {e}")
            import traceback; traceback.print_exc()

        # ③ D-Room ログイン＆物件検索＆Droomシート一括出力
        droom_props = []
        droom_bulk_pdf = ""
        try:
            droom_page = await ctx.new_page()
            droom_shot_dir = shot_dir / "droom"
            droom_shot_dir.mkdir(exist_ok=True)

            dr_logged_in = await login_droom(droom_page, droom_shot_dir)
            if dr_logged_in:
                droom_props = await search_droom(droom_page, area, rent_max, droom_shot_dir)
                print(f"  D-Room: {len(droom_props)}件取得")
                if droom_props:
                    droom_bulk_pdf = await download_droom_bulk_pdf(droom_page, case_dir)
            await droom_page.close()
        except Exception as e:
            print(f"  D-Room検索エラー: {e}")
            import traceback; traceback.print_exc()
            try:
                await droom_page.close()
            except Exception:
                pass

        # ④ レオパレス 物件検索＆PDF取得（ログイン不要）
        lp_props = []
        lp_detail_pdfs = {}   # lp_idx → pdf_path
        try:
            lp_page = await ctx.new_page()
            lp_shot_dir = shot_dir / "leopalace"
            lp_shot_dir.mkdir(exist_ok=True)

            lp_props = await search_leopalace(lp_page, area, rent_max, lp_shot_dir,
                                               work_address=case.get('勤務地住所', ''))
            print(f"  レオパレス: {len(lp_props)}件取得")

            for idx, lp in enumerate(lp_props[:3]):
                dp = await download_leopalace_pdf(
                    lp_page, playwright,
                    lp['detail_url'],
                    f"{case_id}_レオパレス_{lp.get('name', f'物件{idx+1}')}",
                    case_dir)
                if dp:
                    lp_detail_pdfs[idx] = dp
                    print(f"  レオパレスPDF{idx+1}: {dp}")

            await lp_page.close()
        except Exception as e:
            print(f"  レオパレス検索エラー: {e}")
            import traceback; traceback.print_exc()
            try:
                await lp_page.close()
            except Exception:
                pass

        # ⑤ リアブロ ログイン＆物件検索（PDF取得は地図取得後に行う）
        reabro_props = []
        reabro_page = None
        try:
            reabro_page = await ctx.new_page()
            reabro_shot_dir = shot_dir / "reabro"
            reabro_shot_dir.mkdir(exist_ok=True)

            rb_logged_in = await login_reabro(reabro_page, reabro_shot_dir)
            if rb_logged_in:
                reabro_props = await search_reabro(
                    reabro_page, area, rent_max, reabro_shot_dir,
                    work_address=case.get('勤務地住所', '')
                )
                print(f"  リアブロ: {len(reabro_props)}件取得")
            # reabro_page はこの時点では閉じない（後でPDFダウンロードに使用）
        except Exception as e:
            print(f"  リアブロ検索エラー: {e}")
            import traceback; traceback.print_exc()

        # ⑥ 物件を統合（各ソース最大1件ずつ → 計最大5件）
        for idx2, dp in enumerate(droom_props):
            dp['source'] = 'droom'
            dp['droom_idx'] = idx2
        for idx2, lp in enumerate(lp_props):
            lp['source'] = 'leopalace'
            lp['lp_idx'] = idx2
        for idx2, rp in enumerate(reabro_props):
            rp['source'] = 'reabro'
            rp['reabro_idx'] = idx2

        # ⑥-a フィルタリング（家賃下限・築年数）
        print(f"  ── 物件フィルタリング（家賃下限: 上限の30% / 築年数: 20年以内）──")
        atbb_props   = filter_properties(atbb_props,   rent_max)
        hm_props     = filter_properties(hm_props,     rent_max)
        droom_props  = filter_properties(droom_props,  rent_max)
        lp_props     = filter_properties(lp_props,     rent_max)
        reabro_props = filter_properties(reabro_props, rent_max)

        properties = []
        # 各ソースの先頭1件を追加
        for src_list in [atbb_props, hm_props, droom_props, lp_props, reabro_props]:
            if src_list and len(properties) < 5:
                properties.append(src_list[0])
        # 5件未満なら各ソースの2件目以降を補充
        for src_list in [atbb_props, hm_props, droom_props, lp_props, reabro_props]:
            for p in src_list[1:]:
                if len(properties) >= 5:
                    break
                if p not in properties:
                    properties.append(p)

        # 物件が取れなかった場合はモックデータ
        if not properties:
            print("  物件データが取得できなかったためモックデータを使用")
            properties = [
                {"name": f"プレミアムコート{area}", "address": "大阪市北区梅田1-3-1",
                 "rent": "75,000円", "layout": "1K", "area": "26.2㎡",
                 "age": "築4年", "station": "大阪駅 徒歩5分",
                 "madori_path": "", "photo_label": "物件写真"},
                {"name": "ソレイユ堂島", "address": "大阪市北区堂島2-4-12",
                 "rent": "68,000円", "layout": "1K", "area": "22.8㎡",
                 "age": "築7年", "station": "北新地駅 徒歩3分",
                 "madori_path": "", "photo_label": "物件写真"},
                {"name": "グランドアリスタ中之島", "address": "大阪市北区中之島3-6-8",
                 "rent": "79,000円", "layout": "1LDK", "area": "33.1㎡",
                 "age": "築2年", "station": "渡辺橋駅 徒歩2分",
                 "madori_path": "", "photo_label": "物件写真"},
            ]

        # ⑦ 通勤ルート地図（全物件分、最大5件）
        workplace = (
            case.get('勤務地住所') or
            case.get('勤務先住所（転勤先）') or
            case.get('勤務先住所') or ''
        )
        commute_method = case.get('通勤方法', '')
        if not workplace:
            print("  ⚠ 勤務地住所が未設定のため地図取得をスキップします")
        map_shots = {}
        for i, prop in enumerate(properties[:5], 1):
            addr = prop.get("address", "")
            if addr and workplace:
                mp = str(shot_dir / f"map_{i}.png")
                ok = await get_maps_screenshot(page, addr, workplace, mp,
                                               commute_method=commute_method)
                if ok:
                    map_shots[i] = mp

        # ⑧ リアブロ PDF ダウンロード（地図取得後、browser.close() 前に実行）
        reabro_detail_pdfs = {}   # reabro_idx → {"detail": path, "mototsuke": path, "merged": path}
        if reabro_page and reabro_props:
            try:
                for i, prop in enumerate(properties[:5], 1):
                    if prop.get('source') != 'reabro':
                        continue
                    ri = prop.get('reabro_idx', 0)
                    room_id = prop.get('room_id', '')
                    if not room_id:
                        continue
                    map_path = map_shots.get(i, '')
                    # 地図 PNG を先に保存
                    if map_path and os.path.exists(map_path):
                        prop_safe = re.sub(r'[\\/:*?"<>|]', '_', prop.get('name', f'物件{i}'))[:50]
                        map_out = str(case_dir / f"{case_id}_リアプロ_{prop_safe}_地図.png")
                        shutil.copy(map_path, map_out)
                        print(f"  リアブロ地図保存: {map_out}")
                    pdfs = await download_reabro_pdfs(
                        ctx, room_id, f"{case_id}_リアプロ_{prop.get('name', f'物件{ri+1}')}",
                        case_dir, font_name, map_path,
                        commute_method=commute_method, workplace=workplace)
                    reabro_detail_pdfs[ri] = pdfs
            except Exception as e:
                print(f"  リアブロ PDFエラー: {e}")
                import traceback; traceback.print_exc()
        try:
            if reabro_page:
                await reabro_page.close()
        except Exception:
            pass

        await browser.close()

        # ⑨ ATBB・東建・D-Room・レオパレス：地図別途保存 + 地図付きPDF生成
        for i, prop in enumerate(properties[:5], 1):
            source = prop.get('source', 'atbb')
            prop_safe = re.sub(r'[\\/:*?"<>|]', '_', prop.get('name', f'物件{i}'))[:30]

            if source == 'homemate':
                site_label = '東建'
                detail_pdf = hm_detail_pdfs.get(prop.get('hm_idx', -1), '')
            elif source == 'droom':
                # D-Room は一括PDF
                site_label = 'DRoom'
                detail_pdf = droom_bulk_pdf
            elif source == 'leopalace':
                site_label = 'レオパレス'
                detail_pdf = lp_detail_pdfs.get(prop.get('lp_idx', 0), '')
            elif source == 'reabro':
                # リアブロ は download_reabro_pdfs() 内で処理済み
                continue
            else:
                site_label = 'ATBB'
                detail_pdf = atbb_detail_pdfs.get(prop.get('atbb_idx', i - 1), '')

            map_path = map_shots.get(i, '')

            # 地図を別途保存（{case_id}_{site}_{物件名}_地図.png）
            if map_path and os.path.exists(map_path):
                map_out = str(case_dir / f"{case_id}_{site_label}_{prop_safe}_地図.png")
                shutil.copy(map_path, map_out)
                print(f"  地図保存: {map_out}")

            # D-Room の場合は一括PDF＋地図ページ を合成
            if source == 'droom':
                if detail_pdf and os.path.exists(detail_pdf) and map_path and os.path.exists(map_path):
                    merged_out = str(case_dir / f"{case_id}_DRoom_Droomシート_地図付き.pdf")
                    merge_pdf_with_map_image(detail_pdf, map_path, merged_out, font_name,
                                             commute_method=commute_method, workplace=workplace)
                elif detail_pdf and os.path.exists(detail_pdf):
                    dest = str(case_dir / f"{case_id}_DRoom_Droomシート.pdf")
                    if detail_pdf != dest:
                        shutil.copy(detail_pdf, dest)
                    print(f"  D-Room PDF（地図なし）: {dest}")
                continue

            # ATBB・東建・レオパレス：物件詳細PDF + 地図 → 地図付きPDF
            if detail_pdf and os.path.exists(detail_pdf) and map_path and os.path.exists(map_path):
                merged_out = str(case_dir / f"{case_id}_{site_label}_{prop_safe}_地図付き.pdf")
                merge_pdf_with_map_image(detail_pdf, map_path, merged_out, font_name,
                                         commute_method=commute_method, workplace=workplace)
            elif detail_pdf and os.path.exists(detail_pdf):
                dest = str(case_dir / f"{case_id}_{site_label}_{prop_safe}.pdf")
                if detail_pdf != dest:
                    shutil.copy(detail_pdf, dest)
                print(f"  物件詳細PDF(地図なし): {dest}")

        # ⑩ 提案書PDF生成（case_dir 内に保存）
        name = case.get("氏名", f"case{case_num:02d}")
        name_safe = re.sub(r'[\\/:*?"<>|]', '_', str(name))
        pdf_path = case_dir / f"{case_id}_提案書_{name_safe}.pdf"
        generate_pdf(case, properties[:5], map_shots, font_name, pdf_path)
        return pdf_path

    except Exception as e:
        print(f"エラー: {e}")
        await page.screenshot(path=str(shot_dir / "fatal_error.png"))
        await browser.close()
        raise


async def main():
    setup_dirs()
    font_name = register_japanese_font()

    # 「入力データ」フォルダのExcelを検索（テンプレート除く）
    excel_files = [
        f for f in INPUT_DIR.glob("*.xlsx")
        if not f.name.startswith("テンプレート") and not f.name.startswith("~$")
    ]

    if not excel_files:
        print(f"[エラー] 「入力データ」フォルダにExcelファイルが見つかりません。")
        print(f"  → {INPUT_DIR} にExcelファイルをコピーしてから再実行してください。")
        return

    if len(excel_files) > 1:
        print(f"複数のExcelファイルが見つかりました。最新のファイルを使用します。")
        excel_files.sort(key=lambda f: f.stat().st_mtime, reverse=True)

    excel_path = excel_files[0]
    print(f"入力ファイル: {excel_path.name}")
    cases = read_requests(excel_path)
    print(f"{len(cases)}件の依頼を読み込みました")

    async with async_playwright() as pw:
        for i, case in enumerate(cases, 1):
            print(f"\n=== ケース {i}/{len(cases)}: {case.get('氏名', '')} ===")
            pdf = await process_case(pw, case, i, font_name)
            print(f"出力: {pdf}")

    print(f"\n全件完了。出力フォルダ: {OUTPUT_DIR}")


if __name__ == "__main__":
    asyncio.run(main())
