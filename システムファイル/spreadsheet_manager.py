# -*- coding: utf-8 -*-
"""
ヒアリング回答スプレッドシートを読み込み・更新するモジュール。

【役割】
1. ヒアリングフォーム回答.xlsx を読み込む
2. Gmail由来の勤務地データを進捗IDで紐付けてスプレッドシートに書き込む
3. 企業規定書（後日）の家賃上限を紐付ける
4. 各行から「物件検索パラメータ」辞書を生成して返す
"""

import sys
import io
import re
from pathlib import Path
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 都道府県名リスト（住所中の二重プレフィックス検出に使用）
_KNOWN_PREFECTURES = (
    '北海道', '青森県', '岩手県', '宮城県', '秋田県', '山形県', '福島県',
    '茨城県', '栃木県', '群馬県', '埼玉県', '千葉県', '東京都', '神奈川県',
    '新潟県', '富山県', '石川県', '福井県', '山梨県', '長野県', '岐阜県',
    '静岡県', '愛知県', '三重県', '滋賀県', '京都府', '大阪府', '兵庫県',
    '奈良県', '和歌山県', '鳥取県', '島根県', '岡山県', '広島県', '山口県',
    '徳島県', '香川県', '愛媛県', '高知県', '福岡県', '佐賀県', '長崎県',
    '熊本県', '大分県', '宮崎県', '鹿児島県', '沖縄県',
)

# stdout/stderr のUTF-8化（単体実行時のみ。モジュールとして呼ばれる時はスキップ）
if __name__ == '__main__' and hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')


# ══════════════════════════════════════════════════════════
#  ヒアリングフォーム 列マッピング
#  （スプレッドシート「回答転記用」シートの列番号、1始まり）
# ══════════════════════════════════════════════════════════

# ヘッダー行の列名 → 0始まりインデックス
COLUMN_MAP = {
    'タイムスタンプ':          0,
    'メールアドレス':          1,
    '進捗ID':                  2,
    '氏名':                    3,
    '携帯電話番号':            4,
    '現住所':                  5,
    '入居希望日':              6,
    '引越し業者手配':          7,
    '性別':                    8,
    '国籍':                    9,
    '段ボール数':             10,
    '家電':                   11,
    '家具':                   12,
    'その他荷物':             13,
    '駐車場利用':             14,
    '車種':                   15,
    'ナンバープレート':       16,
    '色':                     17,
    '希望社宅種類':           18,
    'レンタル家具希望':       19,
    'レンタル品目':           20,
    '通勤方法':               21,
    '通勤可能時間':           22,
    '現物件状況':             23,
    'レンタル利用状況':       24,
    'ライフライン':           25,
    '入居形態':               26,
    '同居者情報':             27,
    '緊急連絡先':             28,
    'その他要望':             29,
}

# 追記する拡張列（スプレッドシートに存在しない場合に追加）
EXTRA_COLUMNS = [
    '案件ID',           # Gmailから（フォルダ名・ファイル名に使用）
    '勤務地住所',       # Gmailから（就業先住所）
    '最寄り駅',         # Gmailから
    '企業名',           # Gmailから（就業先企業）
    '就業開始日',       # Gmailから
    '家賃上限（円）',    # 依頼上限賃料 or 企業規定書から
    '希望エリア',       # 最寄り駅 or 市区町村名
    '通勤方法_依頼',    # Gmailからの通勤方法（スプレッドシートより優先）
    '入居形態_依頼',    # Gmailからの入居形態（空欄補完）
    '社宅条件',         # 家具付き等
    '検索ステータス',   # 処理状況
]

# 通勤可能時間 → ATBB の時間フィルタ文字列マッピング
COMMUTE_TIME_MAP = {
    '～10分以内': 10,
    '～15分以内': 15,
    '～20分以内': 20,
    '～30分以内': 30,
    '～45分以内': 45,
    '～60分以内': 60,
    '1時間以内':  60,
    '60分以内':   60,
}

# 入居形態 → 推奨間取り
LAYOUT_MAP = {
    '単身':   '1K,1DK,1LDK',
    '家族':   '2LDK,3LDK',
    '2人':    '1LDK,2LDK',
    '夫婦':   '1LDK,2LDK',
    '子供あり': '2LDK,3LDK',
}

# デフォルト家賃上限（規定書がない場合）
DEFAULT_RENT_MAX = 100_000  # 10万円


# ══════════════════════════════════════════════════════════
#  スプレッドシート読み込み
# ══════════════════════════════════════════════════════════

def _find_col_index(headers: list, candidates: list[str]) -> Optional[int]:
    """ヘッダー行から列インデックスを候補名リストで検索する（部分一致）"""
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_str = str(h).strip()
        for c in candidates:
            if c in h_str or h_str in c:
                return i
    return None


def load_hearing_responses(xlsx_path: Path) -> tuple[list[dict], dict]:
    """
    ヒアリングフォーム回答.xlsx を読み込む。

    Returns
    -------
    (rows, col_indices)
    rows        : list of dict  各行データ
    col_indices : dict          列名→インデックスのマッピング（動的検出）
    """
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    # 「回答転記用」シートを優先して探す
    target_sheet = None
    for name in wb.sheetnames:
        if '回答' in name or '転記' in name:
            target_sheet = wb[name]
            break
    if target_sheet is None:
        target_sheet = wb.active

    ws = target_sheet
    raw_headers = [cell.value for cell in ws[1]]
    print(f"  シート「{ws.title}」: {ws.max_row}行 × {ws.max_column}列")

    # 動的列検出
    col_idx = {}
    col_idx['進捗ID']     = _find_col_index(raw_headers, ['進捗ID', '進捗No', '管理番号', 'ID'])
    col_idx['氏名']       = _find_col_index(raw_headers, ['氏名', '名前', '担当者名'])
    col_idx['タイムスタンプ'] = _find_col_index(raw_headers, ['タイムスタンプ', '送信日時', '回答日'])
    col_idx['メールアドレス'] = _find_col_index(raw_headers, ['メールアドレス', 'Email', 'mail'])
    col_idx['携帯電話番号'] = _find_col_index(raw_headers, ['携帯', '電話番号', 'TEL'])
    col_idx['入居希望日'] = _find_col_index(raw_headers, ['入居希望日', '入居日', '希望日'])
    col_idx['引越し業者'] = _find_col_index(raw_headers, ['引越し業者', '引越業者'])
    col_idx['性別']       = _find_col_index(raw_headers, ['性別'])
    col_idx['国籍']       = _find_col_index(raw_headers, ['国籍'])
    col_idx['希望社宅種類'] = _find_col_index(raw_headers, ['希望する社宅', '社宅の種類', '社宅種類'])
    col_idx['通勤方法']   = _find_col_index(raw_headers, ['通勤方法'])
    col_idx['通勤可能時間'] = _find_col_index(raw_headers, ['通勤可能時間', '通勤時間'])
    col_idx['現物件状況'] = _find_col_index(raw_headers, ['現在のお住まい', '物件状況'])
    col_idx['入居形態']   = _find_col_index(raw_headers, ['入居形態'])
    col_idx['同居者情報'] = _find_col_index(raw_headers, ['同居者'])
    col_idx['その他要望'] = _find_col_index(raw_headers, ['その他要望', 'その他'])

    # 拡張列（追記済みの場合は検出、なければNone）
    col_idx['案件ID']       = _find_col_index(raw_headers, ['案件ID', '案件No'])
    col_idx['勤務地住所']   = _find_col_index(raw_headers, ['勤務地住所', '勤務地', '勤務先'])
    col_idx['最寄り駅']     = _find_col_index(raw_headers, ['最寄り駅', '最寄駅'])
    col_idx['家賃上限（円）'] = _find_col_index(raw_headers, ['家賃上限', '上限家賃'])
    col_idx['検索ステータス'] = _find_col_index(raw_headers, ['検索ステータス', 'ステータス'])

    # 全行データをdictで読み込む
    rows = []
    all_vals = list(ws.iter_rows(min_row=2, values_only=True))
    for i, row_vals in enumerate(all_vals):
        if not any(v is not None for v in row_vals):
            continue

        def get(key: str):
            idx = col_idx.get(key)
            if idx is None or idx >= len(row_vals):
                return None
            v = row_vals[idx]
            return str(v).strip() if v is not None else None

        rows.append({
            '_row_num':      i + 2,  # Excel行番号（1始まりヘッダー含む）
            '_raw':          row_vals,
            '進捗ID':        get('進捗ID'),
            '氏名':          get('氏名'),
            'タイムスタンプ': get('タイムスタンプ'),
            'メールアドレス': get('メールアドレス'),
            '携帯電話番号':  get('携帯電話番号'),
            '入居希望日':    get('入居希望日'),
            '引越し業者':    get('引越し業者'),
            '性別':          get('性別'),
            '国籍':          get('国籍'),
            '希望社宅種類':  get('希望社宅種類'),
            '通勤方法':      get('通勤方法'),
            '通勤可能時間':  get('通勤可能時間'),
            '現物件状況':    get('現物件状況'),
            '入居形態':      get('入居形態'),
            '同居者情報':    get('同居者情報'),
            'その他要望':    get('その他要望'),
            # 拡張列（未記入ならNone）
            '案件ID':        get('案件ID'),
            '勤務地住所':    get('勤務地住所'),
            '最寄り駅':      get('最寄り駅'),
            '家賃上限（円）': get('家賃上限（円）'),
            '検索ステータス': get('検索ステータス'),
        })

    print(f"  読み込み完了: {len(rows)}行")
    return rows, col_idx


# ══════════════════════════════════════════════════════════
#  Gmail データの紐付け
# ══════════════════════════════════════════════════════════

def merge_gmail_data(rows: list[dict], shinkoku_map: dict[str, dict]) -> list[dict]:
    """
    進捗IDをキーにGmailの部屋探し依頼情報をスプレッドシート行に紐付ける。

    shinkoku_map のキーは gmail_reader.build_shinkoku_map() の戻り値:
        work_address, nearest_station, work_start_date, company_name,
        move_in_date, occupancy_type, commute_method, housing_condition,
        rent_limit, moving_required, resident_name, remarks
    """
    merged_count = 0
    unmatched_ids = []
    for row in rows:
        sid = row.get('進捗ID')
        if not sid:
            continue
        info = shinkoku_map.get(str(sid).strip())
        if not info:
            unmatched_ids.append(str(sid).strip())
            continue

        # 案件ID（フォルダ名・ファイル名に使用）
        if info.get('case_id'):
            row['案件ID'] = info['case_id']
        # 勤務地住所（最優先）
        if info.get('work_address'):
            row['勤務地住所'] = info['work_address']
            print(f"    進捗ID={sid}: 勤務地住所={info['work_address'][:40]}")
        # 最寄り駅
        if info.get('nearest_station'):
            row['最寄り駅'] = info['nearest_station']
        # 企業名
        if info.get('company_name'):
            row['企業名'] = info['company_name']
        # 就業開始日
        if info.get('work_start_date'):
            row['就業開始日'] = info['work_start_date']
        # 依頼上限賃料（メールに記載があれば最優先）
        if info.get('rent_limit') is not None:
            row['家賃上限（円）'] = str(info['rent_limit'])
        # 通勤方法（メールの値で上書き）
        if info.get('commute_method'):
            row['通勤方法_依頼'] = info['commute_method']
        # 入居形態（メールの値 or スプレッドシートの値）
        if info.get('occupancy_type'):
            row['入居形態_依頼'] = info['occupancy_type']
        elif row.get('入居形態'):
            row['入居形態_依頼'] = row['入居形態']
        # 社宅条件（家具付き等）
        if info.get('housing_condition'):
            row['社宅条件'] = info['housing_condition']
        # 氏名補完（スプレッドシートに名前がない場合）
        if not row.get('氏名') and info.get('resident_name'):
            row['氏名'] = info['resident_name']
        # 入居希望日補完
        if not row.get('入居希望日') and info.get('move_in_date'):
            row['入居希望日'] = info['move_in_date']
        # 現住所（個人の所在地）
        if info.get('current_address'):
            row['現住所'] = info['current_address']

        merged_count += 1

    print(f"  Gmail紐付け: {merged_count}行に依頼情報を補完しました")
    if unmatched_ids:
        # 件数が多い場合は先頭5件だけ表示（ダミーIDが大量にある場合の対策）
        sample = unmatched_ids[:5]
        rest = len(unmatched_ids) - len(sample)
        msg = ', '.join(sample)
        if rest > 0:
            msg += f" 他{rest}件"
        print(f"  ⚠ メール未紐付け（Excel行には進捗IDがあるがメールになし）: {msg}")
    return rows


# ══════════════════════════════════════════════════════════
#  企業規定書の家賃上限紐付け
# ══════════════════════════════════════════════════════════

def _classify_area(work_address: str, area_classification: dict) -> str:
    """
    就業先住所の都道府県名からエリア区分（A地域/B地域/その他）を返す。
    """
    if not work_address:
        return 'その他'
    for area_name, prefs in area_classification.items():
        if area_name.startswith('_'):  # _comment 等のメタキーをスキップ
            continue
        if not isinstance(prefs, list):
            continue
        for pref in prefs:
            if pref in work_address:
                return area_name
    return 'その他'


def merge_company_rules(rows: list[dict], company_rules: dict) -> list[dict]:
    """
    企業規定書から家賃上限を紐付ける。
    優先順位: ①メール記載の依頼上限賃料 → ②規程（エリア+入居形態で算出）

    company_rules の構造:
        {
            'area_classification': {'A地域': [...都道府県...], 'B地域': [...], ...},
            'rent_limits': {
                '単身': {'A地域': 80000, 'B地域': 70000, 'その他': 60000},
                '家族（3人）': {'A地域': 100000, ...},
            },
            'default': {'単身': 80000, '家族': 100000}
        }
    """
    area_cls   = company_rules.get('area_classification', {})
    rent_limits = company_rules.get('rent_limits', {})
    default    = company_rules.get('default', {'単身': DEFAULT_RENT_MAX, '家族': DEFAULT_RENT_MAX})

    for row in rows:
        # ① メール記載の家賃上限が既にあれば使う
        if row.get('家賃上限（円）'):
            continue

        # ② エリア判定
        work_addr = row.get('勤務地住所', '')
        area = _classify_area(work_addr, area_cls)

        # ③ 入居形態の判定
        occupancy = (row.get('入居形態_依頼') or row.get('入居形態') or '単身')
        if '4人' in occupancy or '4人以上' in occupancy:
            occ_key = '家族（4人以上）'
        elif '3人' in occupancy or '家族' in occupancy:
            occ_key = '家族（3人）'
        elif '2人' in occupancy or '夫婦' in occupancy:
            occ_key = '2人入居'
        else:
            occ_key = '単身'

        # ④ 家賃上限を取得
        if occ_key in rent_limits:
            limit = rent_limits[occ_key].get(area, rent_limits[occ_key].get('その他', DEFAULT_RENT_MAX))
        elif '家族' in occ_key or '2人' in occ_key:
            limit = default.get('家族', DEFAULT_RENT_MAX)
        else:
            limit = default.get('単身', DEFAULT_RENT_MAX)

        row['家賃上限（円）'] = str(limit)

    return rows


# ══════════════════════════════════════════════════════════
#  検索パラメータ生成
# ══════════════════════════════════════════════════════════

def _commute_time_minutes(time_str: Optional[str]) -> int:
    """通勤可能時間文字列を分数（int）に変換する"""
    if not time_str:
        return 30
    for k, v in COMMUTE_TIME_MAP.items():
        if k in time_str:
            return v
    # 数値が直接含まれる場合
    m = re.search(r'(\d+)', str(time_str))
    return int(m.group(1)) if m else 30


def _preferred_layout(nyukyo_forma: Optional[str]) -> str:
    """入居形態から推奨間取りを返す"""
    if not nyukyo_forma:
        return ''
    for k, v in LAYOUT_MAP.items():
        if k in nyukyo_forma:
            return v
    return ''


def _extract_city_from_address(address: str) -> str:
    """
    就業先住所から市区町村名を抽出してATBBフリーワード検索用エリアを生成する。

    例:
      '茨城県 かすみがうら市上稲吉2046番地'          → 'かすみがうら市'
      '東京都港区西新橋1-1-1'                       → '港区'
      '大阪府大阪市北区梅田1-1'                     → '大阪市北区'  # 政令市は区まで含める
      '愛知県名古屋市中区三の丸'                     → '名古屋市中区'
      '栃木県宇都宮市馬場通り4-1-1'                 → '宇都宮市'
      '岩手県 岩手県奥州市江刺岩谷堂字松長根52'       → '奥州市'  # 都道府県二重入力に対応
    """
    if not address:
        return ''
    # 都道府県プレフィックスを除去（後続の処理で「都」「道」「府」「県」が誤ヒットしないように）
    addr = re.sub(r'^.+?[都道府県]\s*', '', address)
    if not addr:
        # 住所が都道府県のみの場合
        m = re.search(r'(.+?[都道府県])', address)
        return m.group(1) if m else address[:10]
    # 入力データの不備で都道府県が二重になっている場合（例: "岩手県奥州市..."）に再除去
    # ※ re.match での判定は「宇都宮市」の「都」等で誤検知するため、既知の都道府県名リストで正確に判定する
    for pref in _KNOWN_PREFECTURES:
        if addr.startswith(pref):
            addr = addr[len(pref):].lstrip()
            break
    # 政令市（市の後に区が続く）: 「〇〇市〇〇区」まで含める
    m = re.search(r'(\S+?市)(\S+?区)', addr)
    if m:
        return m.group(1) + m.group(2)
    # 通常の市区町村（市・区・町・村・郡）
    m = re.search(r'(\S+?[市区町村郡])', addr)
    if m:
        return m.group(1)
    return addr[:10]


def _search_area(row: dict) -> str:
    """
    検索エリアを生成する優先順位:
    1. 最寄り駅（Gmail由来）
    2. 勤務地住所から市区町村抽出
    3. 空文字（フリーワード検索なし）
    """
    station = row.get('最寄り駅', '') or ''
    address = row.get('勤務地住所', '') or ''
    if station.strip():
        return station.strip()
    if address.strip():
        city = _extract_city_from_address(address.strip())
        return city if city else address[:20]
    return ''


def generate_search_params(rows: list[dict]) -> list[dict]:
    """
    スプレッドシート行リストから物件検索パラメータ辞書リストを生成する。

    Gmail由来フィールド（入居形態_依頼, 通勤方法_依頼, 社宅条件）がある場合は
    スプレッドシートの値より優先する。

    Returns
    -------
    list of dict: [
        {
            '管理番号':      '001',        # 進捗ID
            '氏名':          '山田太郎',
            '希望エリア':    '梅田駅',     # 最寄り駅 or 市区町村名
            '家賃上限（円）': 80000,
            '希望間取り':    '1K,1DK,1LDK',
            '通勤時間（分）': 30,
            '入居形態':      '単身',
            '社宅種類':      '家具家電付き社宅',
            '入居希望日':    '2025/09/01',
            '通勤方法':      '徒歩',
            ...
        },
        ...
    ]
    """
    params = []
    skipped = 0

    for row in rows:
        sid = row.get('進捗ID')
        # テストデータ（値が'1'だけの行）や進捗IDなし行はスキップ
        if not sid or str(sid).strip() in ('', 'None'):
            skipped += 1
            continue

        # 希望エリア（最寄り駅 or 市区町村）
        area = _search_area(row)
        if not area:
            # エリアなしは全国検索になり無関係な物件が出るためスキップ
            print(f"  ⚠ 進捗ID={sid}: 勤務地・最寄り駅が未取得のため検索対象外（エリアなし）")
            skipped += 1
            continue

        # 家賃上限
        try:
            rent_max = int(float(str(row.get('家賃上限（円）') or DEFAULT_RENT_MAX).replace(',', '')))
        except (ValueError, TypeError):
            rent_max = DEFAULT_RENT_MAX

        # 入居形態: Gmail由来 > スプレッドシート > デフォルト
        nyukyo = (row.get('入居形態_依頼') or row.get('入居形態') or '単身')

        # 通勤方法: Gmail由来 > スプレッドシート
        commute = (row.get('通勤方法_依頼') or row.get('通勤方法') or '')

        # 社宅種類: Gmail由来の社宅条件 > スプレッドシートの希望社宅種類
        syataku_type = (row.get('社宅条件') or row.get('希望社宅種類') or '')

        # 希望エリアを行にも書き戻す（updateSpreadsheet で使用）
        row['希望エリア'] = area

        # フォルダ・ファイル名の識別子：案件ID（メール由来）を優先し、なければ進捗IDで代替
        anken_id = (row.get('案件ID') or '').strip()

        params.append({
            # main.py の process_case が期待するキー名
            '管理番号':       str(sid).strip(),   # 進捗ID（Excelとの紐付けキー）
            '案件ID':         anken_id,            # 案件ID（フォルダ・ファイル名に使用）
            '氏名':           row.get('氏名') or f"進捗ID_{sid}",
            '希望エリア':     area,
            '家賃上限（円）':  rent_max,
            '希望間取り':     _preferred_layout(nyukyo),
            # 拡張情報
            '通勤時間（分）':  _commute_time_minutes(row.get('通勤可能時間')),
            '入居形態':        nyukyo,
            '社宅種類':        syataku_type,
            '入居希望日':      row.get('入居希望日', ''),
            '通勤方法':        commute,
            '性別':            row.get('性別', ''),
            '国籍':            row.get('国籍', ''),
            '勤務地住所':      row.get('勤務地住所', ''),
            '最寄り駅':        row.get('最寄り駅', ''),
            '企業名':          row.get('企業名', ''),
            '就業開始日':      row.get('就業開始日', ''),
            'メールアドレス':  row.get('メールアドレス', ''),
            '携帯電話番号':    row.get('携帯電話番号', ''),
            'その他要望':      row.get('その他要望', ''),
            '現住所':          row.get('現住所', ''),   # 個人の現住所（メール由来）
        })

    print(f"  検索パラメータ生成: {len(params)}件 / スキップ {skipped}行")
    return params


# ══════════════════════════════════════════════════════════
#  スプレッドシートへの書き戻し
# ══════════════════════════════════════════════════════════

def update_spreadsheet(
    xlsx_path: Path,
    rows: list[dict],
    col_idx: dict,
) -> Path:
    """
    勤務地・家賃上限等を書き戻した「完成版」スプレッドシートを保存する。

    Returns
    -------
    保存したファイルパス
    """
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    target_sheet = None
    for name in wb.sheetnames:
        if '回答' in name or '転記' in name:
            target_sheet = wb[name]
            break
    if target_sheet is None:
        target_sheet = wb.active

    ws = target_sheet
    raw_headers = [cell.value for cell in ws[1]]
    max_col = ws.max_column

    # 拡張列を末尾に追加（まだ存在しない場合）
    extra_col_positions = {}
    for col_name in EXTRA_COLUMNS:
        existing = _find_col_index(raw_headers, [col_name])
        if existing is None:
            max_col += 1
            extra_col_positions[col_name] = max_col
            ws.cell(row=1, column=max_col, value=col_name).font = Font(bold=True, color='0070C0')
        else:
            extra_col_positions[col_name] = existing + 1  # openpyxlは1始まり

    # 各行を更新
    row_by_rownum = {r['_row_num']: r for r in rows}
    highlight = PatternFill('solid', start_color='E2EFDA')  # 薄緑

    for row_num, row in row_by_rownum.items():
        # 希望エリアを再計算（_search_area は行の最新値を使う）
        area = row.get('希望エリア') or _search_area(row)

        updates = {
            '案件ID':        row.get('案件ID', ''),
            '勤務地住所':    row.get('勤務地住所', ''),
            '最寄り駅':      row.get('最寄り駅', ''),
            '企業名':        row.get('企業名', ''),
            '就業開始日':    row.get('就業開始日', ''),
            '家賃上限（円）': row.get('家賃上限（円）', ''),
            '希望エリア':    area,
            '通勤方法_依頼': row.get('通勤方法_依頼', ''),
            '入居形態_依頼': row.get('入居形態_依頼', ''),
            '社宅条件':      row.get('社宅条件', ''),
            '検索ステータス': row.get('検索ステータス', ''),
        }
        for col_name, value in updates.items():
            col_pos = extra_col_positions.get(col_name)
            if col_pos and value:
                cell = ws.cell(row=row_num, column=col_pos, value=value)
                cell.fill = highlight

    # 完成版ファイルは「入力データ/完成版/」フォルダに保存（根っこを汚さない）
    kansei_dir = xlsx_path.parent.parent / '入力データ' / '完成版'
    # 入力ファイルが既に「入力データ/」直下にある場合はその中の「完成版/」へ
    if xlsx_path.parent.name == '入力データ':
        kansei_dir = xlsx_path.parent / '完成版'
    kansei_dir.mkdir(parents=True, exist_ok=True)

    # ファイル名（既に「_完成版_*」が付いている場合は除去してから付け直す）
    stem = re.sub(r'_完成版_\d{8}_\d{6}.*$', '', xlsx_path.stem)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = kansei_dir / f"{stem}_完成版_{timestamp}.xlsx"
    wb.save(str(out_path))
    print(f"  スプレッドシート保存: 入力データ/完成版/{out_path.name}")
    return out_path


# ══════════════════════════════════════════════════════════
#  案件データベース（処理済み案件の蓄積）
# ══════════════════════════════════════════════════════════

# データベースの列定義（順序・表示名）
DB_COLUMNS = [
    ('処理日時',             '処理日時'),
    ('案件ID',               '案件ID'),
    ('進捗ID',               '進捗ID'),
    ('氏名',                 '氏名'),
    ('現住所',               '現住所（個人）'),
    ('勤務地住所',           '検索対象住所（転勤先）'),
    ('希望エリア',           '希望エリア'),
    ('企業名',               '企業名'),
    ('就業開始日',           '就業開始日'),
    ('家賃上限（円）',       '家賃上限（円）'),
    ('入居形態_依頼',        '入居形態'),
    ('入居希望日',           '入居希望日'),
    ('通勤方法_依頼',        '通勤方法'),
    ('検索ステータス',       '検索ステータス'),
    # ── 以下は引越し手配用の列（今後追加予定） ──────────
    # ('選定物件名',         '選定物件名'),
    # ('選定物件住所',       '選定物件住所'),
    # ('お客様回答',         'お客様回答'),
    # ('引越し業者',         '引越し業者'),
    # ('引越し日',           '引越し日'),
    # ('手配状況',           '手配状況'),
]


def append_to_case_database(row: dict, db_path: Path) -> None:
    """
    処理済み案件を案件データベース Excel に追記（または更新）する。

    - db_path が存在しない場合は新規作成する。
    - 同じ案件ID（または進捗ID）が既に存在する場合は上書き更新する。
    - 案件ID・進捗ID がどちらもない行はスキップする。

    Parameters
    ----------
    row     : generate_search_params() が返す案件パラメータ辞書
    db_path : データベース Excel のパス
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # ── ファイルのロードまたは新規作成 ──
    if db_path.exists():
        wb = openpyxl.load_workbook(str(db_path))
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '案件データベース'

        # ヘッダー行を作成
        header_fill = PatternFill('solid', start_color='1F4E79')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, (_, display_name) in enumerate(DB_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=display_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # 列幅を設定
        col_widths = {
            1: 18,  # 処理日時
            2: 12,  # 案件ID
            3: 10,  # 進捗ID
            4: 14,  # 氏名
            5: 36,  # 現住所
            6: 36,  # 検索対象住所
            7: 14,  # 希望エリア
            8: 30,  # 企業名
            9: 14,  # 就業開始日
            10: 14, # 家賃上限
            11: 10, # 入居形態
            12: 14, # 入居希望日
            13: 18, # 通勤方法
            14: 12, # 検索ステータス
        }
        for col_num, width in col_widths.items():
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_num)
            ].width = width

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'

    # ── 既存行の検索（案件ID または 進捗ID で重複チェック） ──
    case_id   = str(row.get('案件ID', '') or '').strip()
    shinkoku  = str(row.get('管理番号', '') or '').strip()   # 検索パラメータでは '管理番号'

    target_row_num = None
    for r in ws.iter_rows(min_row=2):
        existing_case_id  = str(r[1].value or '').strip()   # 2列目: 案件ID
        existing_shinkoku = str(r[2].value or '').strip()   # 3列目: 進捗ID
        if (case_id and existing_case_id == case_id) or \
           (shinkoku and existing_shinkoku == shinkoku):
            target_row_num = r[0].row
            break

    # 書き込み先行番号を決定
    is_update = target_row_num is not None
    if not is_update:
        target_row_num = ws.max_row + 1

    # ── 書き込む値を組み立て ──
    area = row.get('希望エリア') or ''

    values = {
        '処理日時':        datetime.now().strftime('%Y/%m/%d %H:%M'),
        '案件ID':          case_id,
        '進捗ID':          shinkoku,
        '氏名':            row.get('氏名', ''),
        '現住所':          row.get('現住所', ''),
        '勤務地住所':      row.get('勤務地住所', ''),
        '希望エリア':      area,
        '企業名':          row.get('企業名', ''),
        '就業開始日':      row.get('就業開始日', ''),
        '家賃上限（円）':  row.get('家賃上限（円）', ''),
        '入居形態_依頼':   row.get('入居形態_依頼', '') or row.get('入居形態', ''),
        '入居希望日':      row.get('入居希望日', ''),
        '通勤方法_依頼':   row.get('通勤方法_依頼', '') or row.get('通勤方法', ''),
        '検索ステータス':  row.get('検索ステータス', '完了'),
    }

    # ── セルに書き込み ──
    row_fill  = PatternFill('solid', start_color='EBF3FB')
    thin      = Side(style='thin', color='CCCCCC')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (key, _) in enumerate(DB_COLUMNS, 1):
        cell = ws.cell(row=target_row_num, column=col_idx, value=values.get(key, ''))
        cell.fill = row_fill
        cell.alignment = Alignment(vertical='center', wrap_text=False)
        cell.border = border

    db_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(db_path))

    action = '更新' if is_update else '追記'
    print(f"  案件DB {action}: {db_path.name}  行{target_row_num}  "
          f"案件ID={case_id}  {row.get('氏名', '')}  エリア={area}")


# ══════════════════════════════════════════════════════════
#  単体テスト用
# ══════════════════════════════════════════════════════════

if __name__ == '__main__':
    test_path = Path(r'C:\Users\shimi\Desktop\syataku.com\株式会社ビーネックステクノロジーズ 社宅利用ヒアリングフォーム  （回答）.xlsx')
    rows, col_idx = load_hearing_responses(test_path)
    print(f"\n総行数: {len(rows)}")
    # 最初の5行を表示
    for r in rows[:5]:
        print(f"  進捗ID={r['進捗ID']}  氏名={r['氏名']}  "
              f"入居形態={r['入居形態']}  社宅種類={r['希望社宅種類']}  "
              f"通勤時間={r['通勤可能時間']}")
