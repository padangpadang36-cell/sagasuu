# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

out_path = Path(__file__).parent.parent / "入力データ" / "テンプレート_依頼一覧.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "依頼一覧"

headers = [
    "氏名",
    "転勤元住所",
    "勤務先住所（転勤先）",
    "希望エリア",
    "入居希望日",
    "希望社宅種類",
    "家賃上限（円）",
    "希望間取り",
    "通勤方法",
    "通勤可能時間",
]

hints = [
    "例：山田 太郎",
    "例：東京都新宿区西新宿2-8-1",
    "例：大阪府大阪市北区梅田3-1-1",
    "例：大阪市北区",
    "例：2026/07/01",
    "例：家具家電なし社宅",
    "例：80000（数字のみ）",
    "例：1K　または　1LDK",
    "例：電車　または　車",
    "例：30分以内",
]

header_fill = PatternFill("solid", fgColor="1a237e")
header_font = Font(bold=True, color="FFFFFF", name="メイリオ", size=10)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

hint_fill = PatternFill("solid", fgColor="E8EAF6")
hint_font = Font(color="3949AB", name="メイリオ", size=9, italic=True)

data_font = Font(name="メイリオ", size=10)
stripe_fill = PatternFill("solid", fgColor="F5F5F5")

ws.row_dimensions[1].height = 28
ws.row_dimensions[2].height = 18

for col, (header, hint) in enumerate(zip(headers, hints), 1):
    h_cell = ws.cell(row=1, column=col, value=header)
    h_cell.fill = header_fill
    h_cell.font = header_font
    h_cell.alignment = header_align
    h_cell.border = border

    hint_cell = ws.cell(row=2, column=col, value=hint)
    hint_cell.fill = hint_fill
    hint_cell.font = hint_font
    hint_cell.alignment = Alignment(horizontal="left", vertical="center")
    hint_cell.border = border

for row in range(3, 23):
    ws.row_dimensions[row].height = 20
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
        if row % 2 == 0:
            cell.fill = stripe_fill

# 家賃列（G列=7）を数値書式
for row in range(3, 23):
    ws.cell(row=row, column=7).number_format = "#,##0"

col_widths = [14, 30, 30, 15, 14, 20, 14, 12, 10, 14]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(out_path)
print(f"テンプレート作成完了: {out_path}")
