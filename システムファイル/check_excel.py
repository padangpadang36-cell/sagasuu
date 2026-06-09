# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\shimi\Desktop\syataku.com\入力データ\依頼一覧_テスト.xlsx', data_only=True)
ws = wb.active
headers = [cell.value for cell in ws[1]]
print('Headers:', headers)
print()
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(row):
        for i, v in enumerate(row):
            print(f'  [{i}] {headers[i]}: {v}')
        print()
